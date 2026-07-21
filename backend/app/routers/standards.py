from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import (
    Boolean,
    Column,
    DateTime,
    Float,
    ForeignKey,
    Integer,
    String,
    UniqueConstraint,
)
from sqlalchemy.orm import Session

from app.db import Base, get_db
from app.routers.auth import get_current_user
from app import models


router = APIRouter(
    prefix="/api/standards",
    tags=["Performance Standards"],
)


class PerformanceStandard(Base):
    __tablename__ = "performance_standards"

    id = Column(Integer, primary_key=True, index=True)

    # Breed standards are global and therefore use company_id = NULL.
    # Company standards are tenant-specific.
    company_id = Column(
        Integer,
        ForeignKey("companies.id"),
        nullable=True,
        index=True,
    )

    standard_code = Column(String(120), nullable=False, index=True)
    standard_name = Column(String(200), nullable=False)
    standard_type = Column(String(20), nullable=False, index=True)

    module = Column(String(50), nullable=False, default="Layers")
    species = Column(String(50), nullable=False, default="Chicken")
    breed = Column(String(120), nullable=True)
    phase = Column(String(50), nullable=True)
    age_week = Column(Integer, nullable=False, index=True)

    lay_rate_hd_pct = Column(Float, nullable=True)
    avg_egg_weight_g = Column(Float, nullable=True)
    body_weight_g = Column(Float, nullable=True)
    body_weight_min_g = Column(Float, nullable=True)
    body_weight_max_g = Column(Float, nullable=True)
    daily_egg_mass_g = Column(Float, nullable=True)
    cumulative_egg_mass_hh_kg = Column(Float, nullable=True)
    eggs_hh = Column(Float, nullable=True)
    feed_min_g_bird_day = Column(Float, nullable=True)
    feed_max_g_bird_day = Column(Float, nullable=True)
    feed_avg_g_bird_day = Column(Float, nullable=True)
    liveability_pct = Column(Float, nullable=True)

    source_file = Column(String(255), nullable=True)
    active = Column(Boolean, nullable=False, default=True)

    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)
    updated_at = Column(
        DateTime,
        nullable=False,
        default=datetime.utcnow,
        onupdate=datetime.utcnow,
    )
    imported_by = Column(String(200), nullable=True)

    __table_args__ = (
        UniqueConstraint(
            "standard_code",
            "standard_type",
            "company_id",
            "age_week",
            name="uq_performance_standard_scope_age",
        ),
    )


IMPORT_SHEET = "Standards_Import"

REQUIRED_COLUMNS = {
    "standard_code",
    "standard_name",
    "standard_type",
    "module",
    "species",
    "breed",
    "phase",
    "age_week",
    "active",
}

FLOAT_COLUMNS = [
    "lay_rate_hd_pct",
    "avg_egg_weight_g",
    "body_weight_g",
    "body_weight_min_g",
    "body_weight_max_g",
    "daily_egg_mass_g",
    "cumulative_egg_mass_hh_kg",
    "eggs_hh",
    "feed_min_g_bird_day",
    "feed_max_g_bird_day",
    "feed_avg_g_bird_day",
    "liveability_pct",
]

ALL_COLUMNS = [
    "standard_code",
    "standard_name",
    "standard_type",
    "module",
    "species",
    "breed",
    "phase",
    "age_week",
    *FLOAT_COLUMNS,
    "source_file",
    "active",
]


def _require_admin(current_user: models.AppUser) -> None:
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )


def _resolve_scope_company_id(
    current_user: models.AppUser,
    standard_type: str,
    requested_company_id: int | None,
) -> int | None:
    normalised_type = standard_type.strip().title()

    if normalised_type not in {"Breed", "Company"}:
        raise HTTPException(
            status_code=400,
            detail="standard_type must be Breed or Company",
        )

    if normalised_type == "Breed":
        if not current_user.is_global_admin:
            raise HTTPException(
                status_code=403,
                detail="Only Global Admin may import global breed standards",
            )
        return None

    if current_user.is_global_admin:
        if requested_company_id is None:
            raise HTTPException(
                status_code=400,
                detail="company_id is required for a Company standard",
            )
        return requested_company_id

    if current_user.company_id is None:
        raise HTTPException(
            status_code=403,
            detail="Your account is not assigned to a company",
        )

    return current_user.company_id


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Expected a number but received {value!r}")


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    if value is None or value == "":
        return True

    normalised = str(value).strip().lower()

    if normalised in {"1", "true", "yes", "y", "active"}:
        return True

    if normalised in {"0", "false", "no", "n", "inactive"}:
        return False

    raise ValueError(f"Expected TRUE/FALSE but received {value!r}")


def _normalise_row(
    raw: dict[str, Any],
    forced_standard_type: str,
    source_filename: str,
) -> dict[str, Any]:
    standard_code = str(raw.get("standard_code") or "").strip()
    standard_name = str(raw.get("standard_name") or "").strip()

    if not standard_code:
        raise ValueError("standard_code is required")

    if not standard_name:
        raise ValueError("standard_name is required")

    try:
        age_week = int(raw.get("age_week"))
    except (TypeError, ValueError):
        raise ValueError("age_week must be a whole number")

    if age_week < 0 or age_week > 200:
        raise ValueError("age_week must be between 0 and 200")

    output: dict[str, Any] = {
        "standard_code": standard_code.upper().replace(" ", "_"),
        "standard_name": standard_name,
        "standard_type": forced_standard_type,
        "module": str(raw.get("module") or "Layers").strip() or "Layers",
        "species": str(raw.get("species") or "Chicken").strip() or "Chicken",
        "breed": str(raw.get("breed") or "").strip() or None,
        "phase": str(raw.get("phase") or "").strip() or None,
        "age_week": age_week,
        "source_file": (
            str(raw.get("source_file") or "").strip()
            or source_filename
        ),
        "active": _to_bool(raw.get("active")),
    }

    for column_name in FLOAT_COLUMNS:
        output[column_name] = _to_float(raw.get(column_name))

    return output


def _read_import_rows(
    file_bytes: bytes,
    source_filename: str,
    forced_standard_type: str,
) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=f"Could not read the Excel workbook: {error}",
        )

    if IMPORT_SHEET not in workbook.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=(
                f"The workbook must contain a worksheet named "
                f"{IMPORT_SHEET!r}."
            ),
        )

    worksheet = workbook[IMPORT_SHEET]
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(
            status_code=400,
            detail="The import worksheet is empty",
        )

    headers = [
        str(value).strip() if value is not None else ""
        for value in header_row
    ]

    missing = sorted(REQUIRED_COLUMNS.difference(headers))

    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing required import columns: "
                + ", ".join(missing)
            ),
        )

    parsed_rows: list[dict[str, Any]] = []
    errors: list[str] = []

    for excel_row_number, values in enumerate(rows, start=2):
        raw = dict(zip(headers, values))

        if not any(
            value is not None and str(value).strip() != ""
            for value in values
        ):
            continue

        try:
            parsed_rows.append(
                _normalise_row(
                    raw,
                    forced_standard_type=forced_standard_type,
                    source_filename=source_filename,
                )
            )
        except ValueError as error:
            errors.append(
                f"Row {excel_row_number}: {error}"
            )

    if errors:
        preview = errors[:20]
        remaining = len(errors) - len(preview)

        detail = "\n".join(preview)

        if remaining > 0:
            detail += f"\n...and {remaining} more row errors."

        raise HTTPException(
            status_code=400,
            detail=detail,
        )

    if not parsed_rows:
        raise HTTPException(
            status_code=400,
            detail="No standard rows were found to import",
        )

    codes = {row["standard_code"] for row in parsed_rows}

    if len(codes) != 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "Each uploaded workbook must contain exactly one "
                "standard_code."
            ),
        )

    duplicate_ages = set()
    seen_ages = set()

    for row in parsed_rows:
        age_week = row["age_week"]

        if age_week in seen_ages:
            duplicate_ages.add(age_week)

        seen_ages.add(age_week)

    if duplicate_ages:
        raise HTTPException(
            status_code=400,
            detail=(
                "Duplicate age_week values in the workbook: "
                + ", ".join(str(value) for value in sorted(duplicate_ages))
            ),
        )

    return parsed_rows


def _serialise_standard(row: PerformanceStandard) -> dict[str, Any]:
    return {
        "id": row.id,
        "company_id": row.company_id,
        "standard_code": row.standard_code,
        "standard_name": row.standard_name,
        "standard_type": row.standard_type,
        "module": row.module,
        "species": row.species,
        "breed": row.breed,
        "phase": row.phase,
        "age_week": row.age_week,
        "lay_rate_hd_pct": row.lay_rate_hd_pct,
        "avg_egg_weight_g": row.avg_egg_weight_g,
        "body_weight_g": row.body_weight_g,
        "body_weight_min_g": row.body_weight_min_g,
        "body_weight_max_g": row.body_weight_max_g,
        "daily_egg_mass_g": row.daily_egg_mass_g,
        "cumulative_egg_mass_hh_kg": row.cumulative_egg_mass_hh_kg,
        "eggs_hh": row.eggs_hh,
        "feed_min_g_bird_day": row.feed_min_g_bird_day,
        "feed_max_g_bird_day": row.feed_max_g_bird_day,
        "feed_avg_g_bird_day": row.feed_avg_g_bird_day,
        "liveability_pct": row.liveability_pct,
        "source_file": row.source_file,
        "active": row.active,
        "created_at": row.created_at,
        "updated_at": row.updated_at,
        "imported_by": row.imported_by,
    }


@router.get("")
def list_standards(
    company_id: int | None = None,
    include_rows: bool = False,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(PerformanceStandard)

    if current_user.is_global_admin:
        if company_id is not None:
            query = query.filter(
                (PerformanceStandard.company_id.is_(None))
                | (PerformanceStandard.company_id == company_id)
            )
    else:
        query = query.filter(
            (PerformanceStandard.company_id.is_(None))
            | (
                PerformanceStandard.company_id
                == current_user.company_id
            )
        )

    rows = (
        query
        .order_by(
            PerformanceStandard.standard_name.asc(),
            PerformanceStandard.age_week.asc(),
        )
        .all()
    )

    if include_rows:
        return [_serialise_standard(row) for row in rows]

    grouped: dict[tuple[str, str, int | None], dict[str, Any]] = {}

    for row in rows:
        key = (
            row.standard_code,
            row.standard_type,
            row.company_id,
        )

        item = grouped.setdefault(
            key,
            {
                "standard_code": row.standard_code,
                "standard_name": row.standard_name,
                "standard_type": row.standard_type,
                "company_id": row.company_id,
                "module": row.module,
                "species": row.species,
                "breed": row.breed,
                "active": row.active,
                "age_min": row.age_week,
                "age_max": row.age_week,
                "row_count": 0,
                "updated_at": row.updated_at,
                "source_file": row.source_file,
            },
        )

        item["row_count"] += 1
        item["age_min"] = min(item["age_min"], row.age_week)
        item["age_max"] = max(item["age_max"], row.age_week)
        item["active"] = item["active"] or row.active

        if row.updated_at and (
            item["updated_at"] is None
            or row.updated_at > item["updated_at"]
        ):
            item["updated_at"] = row.updated_at

    return list(grouped.values())


@router.get("/{standard_code}")
def get_standard_rows(
    standard_code: str,
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(PerformanceStandard).filter(
        PerformanceStandard.standard_code
        == standard_code.strip().upper()
    )

    if current_user.is_global_admin:
        if company_id is None:
            query = query.filter(
                PerformanceStandard.company_id.is_(None)
            )
        else:
            query = query.filter(
                PerformanceStandard.company_id == company_id
            )
    else:
        query = query.filter(
            (PerformanceStandard.company_id.is_(None))
            | (
                PerformanceStandard.company_id
                == current_user.company_id
            )
        )

    rows = (
        query
        .order_by(PerformanceStandard.age_week.asc())
        .all()
    )

    if not rows:
        raise HTTPException(
            status_code=404,
            detail="Standard not found",
        )

    return [_serialise_standard(row) for row in rows]


@router.post("/import")
async def import_standard(
    file: UploadFile = File(...),
    standard_type: str = Form(...),
    company_id: int | None = Form(default=None),
    replace_existing: bool = Form(default=True),
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)

    normalised_type = standard_type.strip().title()
    scope_company_id = _resolve_scope_company_id(
        current_user,
        normalised_type,
        company_id,
    )

    filename = file.filename or "standards.xlsx"

    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx files are supported",
        )

    file_bytes = await file.read()

    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail="The standards workbook is larger than 10 MB",
        )

    parsed_rows = _read_import_rows(
        file_bytes=file_bytes,
        source_filename=filename,
        forced_standard_type=normalised_type,
    )

    standard_code = parsed_rows[0]["standard_code"]

    existing_query = db.query(PerformanceStandard).filter(
        PerformanceStandard.standard_code == standard_code,
        PerformanceStandard.standard_type == normalised_type,
    )

    if scope_company_id is None:
        existing_query = existing_query.filter(
            PerformanceStandard.company_id.is_(None)
        )
    else:
        existing_query = existing_query.filter(
            PerformanceStandard.company_id == scope_company_id
        )

    existing_count = existing_query.count()

    if existing_count > 0 and not replace_existing:
        raise HTTPException(
            status_code=409,
            detail=(
                "This standard already exists. Enable replace_existing "
                "to replace the current rows."
            ),
        )

    try:
        if existing_count > 0:
            existing_query.delete(synchronize_session=False)

        now = datetime.utcnow()

        for row in parsed_rows:
            db.add(
                PerformanceStandard(
                    company_id=scope_company_id,
                    imported_by=current_user.full_name,
                    created_at=now,
                    updated_at=now,
                    **row,
                )
            )

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "standard_code": standard_code,
        "standard_name": parsed_rows[0]["standard_name"],
        "standard_type": normalised_type,
        "company_id": scope_company_id,
        "rows_imported": len(parsed_rows),
        "rows_replaced": existing_count,
        "source_file": filename,
    }


@router.patch("/{standard_code}/active")
def set_standard_active(
    standard_code: str,
    active: bool,
    company_id: int | None = None,
    standard_type: str = "Breed",
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)

    normalised_type = standard_type.strip().title()
    scope_company_id = _resolve_scope_company_id(
        current_user,
        normalised_type,
        company_id,
    )

    query = db.query(PerformanceStandard).filter(
        PerformanceStandard.standard_code
        == standard_code.strip().upper(),
        PerformanceStandard.standard_type == normalised_type,
    )

    if scope_company_id is None:
        query = query.filter(
            PerformanceStandard.company_id.is_(None)
        )
    else:
        query = query.filter(
            PerformanceStandard.company_id == scope_company_id
        )

    rows = query.all()

    if not rows:
        raise HTTPException(
            status_code=404,
            detail="Standard not found",
        )

    now = datetime.utcnow()

    for row in rows:
        row.active = active
        row.updated_at = now

    db.commit()

    return {
        "ok": True,
        "standard_code": standard_code.strip().upper(),
        "active": active,
        "rows_updated": len(rows),
    }


@router.delete("/{standard_code}")
def delete_standard(
    standard_code: str,
    company_id: int | None = None,
    standard_type: str = "Breed",
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)

    normalised_type = standard_type.strip().title()
    scope_company_id = _resolve_scope_company_id(
        current_user,
        normalised_type,
        company_id,
    )

    query = db.query(PerformanceStandard).filter(
        PerformanceStandard.standard_code
        == standard_code.strip().upper(),
        PerformanceStandard.standard_type == normalised_type,
    )

    if scope_company_id is None:
        query = query.filter(
            PerformanceStandard.company_id.is_(None)
        )
    else:
        query = query.filter(
            PerformanceStandard.company_id == scope_company_id
        )

    deleted = query.delete(synchronize_session=False)

    if deleted == 0:
        raise HTTPException(
            status_code=404,
            detail="Standard not found",
        )

    db.commit()

    return {
        "ok": True,
        "standard_code": standard_code.strip().upper(),
        "rows_deleted": deleted,
    }
