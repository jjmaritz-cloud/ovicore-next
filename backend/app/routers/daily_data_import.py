from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app import models
from app.db import get_db
from app.routers.auth import get_current_user

router = APIRouter(prefix="/api/admin", tags=["Daily Data Import"])

MODULES: dict[str, dict[str, Any]] = {
    "broilers": {
        "label": "Broilers",
        "flock_model": models.BroilerPlacementPlan,
        "daily_model": models.BroilerDailyPerformance,
        "flock_code": "cycle_code",
        "fk": "placement_plan_id",
        "start_date": "placement_date",
        "opening": "planned_birds",
        "fields": [
            "Mortality Front", "Mortality Middle", "Mortality Back", "Mortality Other",
            "Cull Legs", "Cull Runts", "Cull Beak", "Cull Other",
            "Feed kg", "Water L", "Bodyweight kg", "Notes",
        ],
    },
    "breeder_rearing": {
        "label": "Breeder Rearing",
        "flock_model": models.BreederRearingFlock,
        "daily_model": models.BreederRearingDailyPerformance,
        "flock_code": "flock_code",
        "fk": "flock_id",
        "start_date": "hatch_date",
        "opening": None,
        "fields": [
            "Opening Female Birds", "Female Mortality", "Female Culls",
            "Opening Male Birds", "Male Mortality", "Male Culls",
            "Feed kg", "Water L", "Female Bodyweight kg", "Male Bodyweight kg", "Notes",
        ],
    },
    "breeder_production": {
        "label": "Breeder Production",
        "flock_model": models.BreederProductionFlock,
        "daily_model": models.BreederProductionDailyPerformance,
        "flock_code": "flock_code",
        "fk": "flock_id",
        "start_date": "hatch_date",
        "opening": None,
        "fields": [
            "Opening Female Birds", "Female Mortality", "Female Culls",
            "Opening Male Birds", "Male Mortality", "Male Culls",
            "Feed kg", "Water L", "Female Bodyweight kg", "Male Bodyweight kg",
            "Total Eggs", "Hatching Eggs", "Floor Eggs", "Rejects",
            "Production Standard %", "Notes",
        ],
    },
    "commercial_rearing": {
        "label": "Commercial Rearing",
        "flock_model": models.LayerRearingFlock,
        "daily_model": models.LayerRearingDailyPerformance,
        "flock_code": "flock_code",
        "fk": "flock_id",
        "start_date": "hatch_date",
        "opening": "birds_placed",
        "fields": [
            "Mortality Front", "Mortality Middle", "Mortality Back", "Mortality Other",
            "Cull Legs", "Cull Runts", "Cull Beak", "Cull Other",
            "Feed kg", "Water L", "Bodyweight kg", "Notes",
        ],
    },
    "commercial_layers": {
        "label": "Commercial Layers",
        "flock_model": models.CommercialLayerFlock,
        "daily_model": models.CommercialLayerDailyPerformance,
        "flock_code": "flock_code",
        "fk": "flock_id",
        "start_date": "hatch_date",
        "opening": "birds_housed",
        "fields": [
            "Mortality Birds", "Cull Birds", "Total Eggs", "Egg Weight g",
            "Feed kg", "Water L", "Bodyweight g", "Production Standard %",
            "Mortality Standard %", "Egg Weight Standard g", "Feed Standard g/bird/day",
            "Eggs per Bird Standard", "Bodyweight Standard g", "Notes",
        ],
    },
}

COMMON_HEADERS = ["Flock Code", "Entry Date", "Age Days"]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _date(value: Any, row: int, errors: list[str]) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    errors.append(f"Row {row}: Entry Date is invalid.")
    return None


def _number(value: Any, name: str, row: int, errors: list[str], integer: bool = False):
    if value is None or _text(value) == "":
        return None
    try:
        number = float(value)
        return int(number) if integer else number
    except (TypeError, ValueError):
        errors.append(f"Row {row}: {name} must be numeric.")
        return None


def _records(workbook) -> list[tuple[int, dict[str, Any]]]:
    if "Daily Data" not in workbook.sheetnames:
        return []
    sheet = workbook["Daily Data"]
    headers = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells]
        if not any(value is not None and _text(value) for value in values):
            continue
        rows.append((row_number, {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}))
    return rows


def _module_enabled(company, module: str) -> bool:
    if module in {"broilers"}:
        return bool(company.enable_broilers)
    if module in {"breeder_rearing", "breeder_production"}:
        return bool(company.enable_breeders)
    if module in {"commercial_rearing", "commercial_layers"}:
        return bool(company.enable_layers)
    return False


def _require_admin(user):
    if not (user.is_global_admin or user.is_company_admin):
        raise HTTPException(status_code=403, detail="Admin access required")


def _resolve_company(db: Session, user, company_id: int):
    _require_admin(user)
    if not user.is_global_admin and user.company_id != company_id:
        raise HTTPException(status_code=403, detail="You do not have access to this company")
    company = db.query(models.Company).filter(models.Company.id == company_id, models.Company.active == True).first()
    if not company:
        raise HTTPException(status_code=404, detail="Selected company was not found or is inactive")
    return company


@router.get("/daily-data-template/{module}")
def download_daily_template(module: str):
    config = MODULES.get(module)
    if not config:
        raise HTTPException(status_code=404, detail="Unknown production module")
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Data"
    headers = COMMON_HEADERS + config["fields"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B5D4B")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = max(14, min(28, len(header) + 3))
    ws.append(["EXAMPLE-FLOCK", date.today(), 0] + [None] * len(config["fields"]))
    instructions = wb.create_sheet("Instructions")
    instructions.append([f"OviCore {config['label']} Daily House Card Import"])
    instructions.append(["Keep the Daily Data sheet name and header names unchanged."])
    instructions.append(["Calculated fields such as closing birds, totals, livability and production percentages are calculated by OviCore."])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"OviCore_{module}_daily_data_template.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _set_common_growout(entry, row, row_number, errors):
    int_map = {
        "mortality_front": "Mortality Front", "mortality_middle": "Mortality Middle",
        "mortality_back": "Mortality Back", "mortality_other": "Mortality Other",
        "cull_legs": "Cull Legs", "cull_runts": "Cull Runts",
        "cull_beak": "Cull Beak", "cull_other": "Cull Other",
    }
    for attr, header in int_map.items():
        setattr(entry, attr, _number(row.get(header), header, row_number, errors, True) or 0)
    entry.mortality_birds = sum(int(getattr(entry, key) or 0) for key in ("mortality_front", "mortality_middle", "mortality_back", "mortality_other"))
    entry.cull_birds = sum(int(getattr(entry, key) or 0) for key in ("cull_legs", "cull_runts", "cull_beak", "cull_other"))
    entry.feed_kg = _number(row.get("Feed kg"), "Feed kg", row_number, errors)
    entry.water_litres = _number(row.get("Water L"), "Water L", row_number, errors)
    entry.avg_weight_kg = _number(row.get("Bodyweight kg"), "Bodyweight kg", row_number, errors)
    entry.notes = _text(row.get("Notes"))


def _set_breeder(entry, row, row_number, errors, production: bool):
    for attr, header in {
        "opening_female_birds": "Opening Female Birds", "female_mortality": "Female Mortality",
        "female_culls": "Female Culls", "opening_male_birds": "Opening Male Birds",
        "male_mortality": "Male Mortality", "male_culls": "Male Culls",
    }.items():
        value = _number(row.get(header), header, row_number, errors, True)
        if value is not None or attr not in {"opening_female_birds", "opening_male_birds"}:
            setattr(entry, attr, value or 0)
    entry.feed_kg = _number(row.get("Feed kg"), "Feed kg", row_number, errors)
    entry.water_litres = _number(row.get("Water L"), "Water L", row_number, errors)
    entry.female_bodyweight_kg = _number(row.get("Female Bodyweight kg"), "Female Bodyweight kg", row_number, errors)
    entry.male_bodyweight_kg = _number(row.get("Male Bodyweight kg"), "Male Bodyweight kg", row_number, errors)
    if production:
        for attr, header in {"total_eggs":"Total Eggs", "hatching_eggs":"Hatching Eggs", "floor_eggs":"Floor Eggs", "rejects":"Rejects"}.items():
            setattr(entry, attr, _number(row.get(header), header, row_number, errors, True) or 0)
        entry.production_standard_pct = _number(row.get("Production Standard %"), "Production Standard %", row_number, errors)
    entry.notes = _text(row.get("Notes"))


def _set_layers(entry, row, row_number, errors):
    for attr, header in {"mortality_birds":"Mortality Birds", "cull_birds":"Cull Birds", "total_eggs":"Total Eggs"}.items():
        setattr(entry, attr, _number(row.get(header), header, row_number, errors, True) or 0)
    for attr, header in {
        "egg_weight_g":"Egg Weight g", "feed_kg":"Feed kg", "water_litres":"Water L", "bodyweight_g":"Bodyweight g",
        "production_standard_pct":"Production Standard %", "mortality_standard_pct":"Mortality Standard %",
        "egg_weight_standard_g":"Egg Weight Standard g", "feed_standard_g_bird_day":"Feed Standard g/bird/day",
        "eggs_per_bird_standard":"Eggs per Bird Standard", "bodyweight_standard_g":"Bodyweight Standard g",
    }.items():
        setattr(entry, attr, _number(row.get(header), header, row_number, errors))
    entry.notes = _text(row.get("Notes"))


def _recalculate_sequence(db: Session, module: str, flock, user_name: str):
    config = MODULES[module]
    daily = config["daily_model"]
    fk = config["fk"]
    rows = db.query(daily).filter(getattr(daily, fk) == flock.id).order_by(daily.entry_date.asc(), daily.id.asc()).all()
    if module in {"broilers", "commercial_rearing", "commercial_layers"}:
        opening = int(getattr(flock, config["opening"], 0) or 0)
        for entry in rows:
            entry.opening_birds = opening
            if module in {"broilers", "commercial_rearing"}:
                entry.mortality_birds = sum(int(getattr(entry, k) or 0) for k in ("mortality_front","mortality_middle","mortality_back","mortality_other"))
                entry.cull_birds = sum(int(getattr(entry, k) or 0) for k in ("cull_legs","cull_runts","cull_beak","cull_other"))
            entry.closing_birds = max(0, opening - int(entry.mortality_birds or 0) - int(entry.cull_birds or 0))
            opening = entry.closing_birds
            entry.last_saved_by = user_name
            entry.last_saved_at = datetime.utcnow()
    else:
        female = int(getattr(flock, "opening_female_birds", None) or getattr(flock, "female_birds", 0) or 0)
        male = int(getattr(flock, "opening_male_birds", None) or getattr(flock, "male_birds", 0) or 0)
        for entry in rows:
            entry.opening_female_birds = female
            entry.opening_male_birds = male
            entry.closing_female_birds = max(0, female - int(entry.female_mortality or 0) - int(entry.female_culls or 0))
            entry.closing_male_birds = max(0, male - int(entry.male_mortality or 0) - int(entry.male_culls or 0))
            female, male = entry.closing_female_birds, entry.closing_male_birds
            entry.last_saved_by = user_name
            entry.last_saved_at = datetime.utcnow()


@router.post("/daily-data-import")
async def import_daily_data(
    company_id: int = Form(...), module: str = Form(...), commit: bool = Form(False),
    allow_updates: bool = Form(False), workbook: UploadFile = File(...),
    current_user: models.AppUser = Depends(get_current_user), db: Session = Depends(get_db),
):
    company = _resolve_company(db, current_user, company_id)
    config = MODULES.get(module)
    if not config:
        raise HTTPException(status_code=400, detail="Select a valid production module")
    if not _module_enabled(company, module):
        raise HTTPException(status_code=400, detail=f"{config['label']} is not enabled for this company")
    filename = workbook.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx workbook")
    try:
        excel = load_workbook(BytesIO(await workbook.read()), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {exc}")
    records = _records(excel)
    errors: list[str] = []
    warnings: list[str] = []
    if not records:
        errors.append("The Daily Data sheet is missing or contains no data rows.")
    flock_model = config["flock_model"]
    daily_model = config["daily_model"]
    flock_rows = db.query(flock_model).filter(flock_model.company_id == company_id).all()
    flock_by_code = {_text(getattr(f, config["flock_code"])).lower(): f for f in flock_rows}
    existing_rows = db.query(daily_model).filter(daily_model.company_id == company_id).all()
    existing_by_key = {(getattr(e, config["fk"]), e.entry_date): e for e in existing_rows}
    parsed = []
    seen = set()
    counts = {"create": 0, "update": 0, "unchanged": 0}
    for row_number, row in records:
        code = _text(row.get("Flock Code"))
        entry_date = _date(row.get("Entry Date"), row_number, errors)
        if not code or not entry_date:
            if not code: errors.append(f"Row {row_number}: Flock Code is required.")
            continue
        flock = flock_by_code.get(code.lower())
        if not flock:
            errors.append(f"Row {row_number}: Flock Code '{code}' was not found in {config['label']}.")
            continue
        key = (flock.id, entry_date)
        if key in seen:
            errors.append(f"Row {row_number}: duplicate entry for '{code}' on {entry_date}.")
            continue
        seen.add(key)
        age_days = _number(row.get("Age Days"), "Age Days", row_number, errors, True)
        start_date = getattr(flock, config["start_date"], None)
        if age_days is None and start_date:
            age_days = (entry_date - start_date).days
        if start_date and entry_date < start_date:
            errors.append(f"Row {row_number}: Entry Date is before the flock start/hatch date.")
        existing = existing_by_key.get(key)
        counts["create" if existing is None else "update" if allow_updates else "unchanged"] += 1
        parsed.append((row_number, row, flock, entry_date, age_days, existing))
    result = {"company":{"id":company.id,"name":company.company_name}, "module":module, "module_label":config["label"], "filename":filename, "mode":"commit" if commit else "preview", "allow_updates":allow_updates, "committed":False, "actions":{"performance":counts}, "errors":errors, "warnings":warnings, "rows":len(records)}
    if errors or not commit:
        return result
    affected = set()
    try:
        for row_number, row, flock, entry_date, age_days, existing in parsed:
            if existing is not None and not allow_updates:
                continue
            entry = existing or daily_model(company_id=company_id, **{config["fk"]: flock.id}, entry_date=entry_date)
            entry.entry_date = entry_date
            entry.age_days = age_days
            if module in {"broilers", "commercial_rearing"}:
                _set_common_growout(entry, row, row_number, errors)
            elif module in {"breeder_rearing", "breeder_production"}:
                _set_breeder(entry, row, row_number, errors, module == "breeder_production")
            else:
                _set_layers(entry, row, row_number, errors)
            entry.last_saved_by = current_user.full_name
            entry.last_saved_at = datetime.utcnow()
            if existing is None: db.add(entry)
            affected.add(flock.id)
        if errors:
            db.rollback()
            result["errors"] = errors
            return result
        db.flush()
        for flock_id in affected:
            flock = next(f for f in flock_rows if f.id == flock_id)
            _recalculate_sequence(db, module, flock, current_user.full_name)
        db.commit()
        result["committed"] = True
        return result
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Daily import failed and was rolled back: {exc}")
