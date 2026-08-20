import os
import base64
import json
import re
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from io import BytesIO
from fastapi import Body, Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import inspect, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload
from openpyxl import load_workbook
from app.routers import broiler_processing
from app.routers import app_notes
from app.routers import broiler_supply
from app.routers import hatchery
from app.routers import access
from app.routers import auth
from app.routers import standards
from app.routers import daily_data_import
from app.routers.standards import PerformanceStandard
from app.routers.auth import get_current_user
from app import models

from .db import Base, engine, SessionLocal, get_db
from .models import (
    Company,
    BroilerFarm,
    BroilerShed,
    BroilerPlacementPlan,
    BroilerDailyPerformance,
    BroilerPaperCapture,
)
from .schemas import (
    BroilerDemandPlanCreate,
    BroilerDemandPlanOut,
    BroilerDemandPlanPatch,
    BroilerFarmCreate,
    BroilerFarmPatch,
    BroilerFarmOut,
    BroilerShedCreate,
    BroilerShedPatch,
    BroilerShedOut,
    BroilerDailyPerformanceCreate,
    BroilerDailyPerformancePatch,
    BroilerDailyPerformanceOut,
    BroilerPaperCaptureSourceData,
    BroilerPaperCaptureReview,
    BroilerPaperCaptureExtractOut,
    BroilerPaperCaptureApproveIn,
    BroilerPaperCaptureApproveOut,
    LayerRearingFlockCreate,
    LayerRearingFlockPatch,
    LayerRearingFlockOut,
    BreederRearingFlockCreate,
    BreederRearingFlockPatch,
    BreederRearingFlockOut,
    BreederRearingTransferCreate,
    BreederProductionFlockOut,
    BreederTransferResult,
    BreederProductionDailyPerformanceCreate,
    BreederProductionDailyPerformancePatch,
    BreederProductionDailyPerformanceOut,
    CommercialLayerFlockCreate,
    CommercialLayerFlockPatch,
    CommercialLayerFlockOut,
    CommercialLayerDailyPerformanceCreate,
    CommercialLayerDailyPerformancePatch,
    CommercialLayerPerformanceOut,
    LayerRearingTransferCreate,
    LayerRearingTransferResult,
)
from .calculations import build_plan_response
from .seed import seed_demo_data

app = FastAPI(title="OviCore Broiler Module API", version="0.1.0")
app.include_router(broiler_processing.router)
app.include_router(app_notes.router)
app.include_router(broiler_supply.router)
app.include_router(hatchery.router)
app.include_router(access.router)
app.include_router(auth.router)
app.include_router(standards.router)
app.include_router(daily_data_import.router)

origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:3001",
    "http://127.0.0.1:3001",
]

env_origins = os.getenv("CORS_ORIGINS")
if env_origins:
    origins.extend([origin.strip() for origin in env_origins.split(",") if origin.strip()])

app.add_middleware(
    CORSMiddleware,
    allow_origins=list(set(origins)),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def resolve_company_id(
    current_user: models.AppUser,
    requested_company_id: int | None = None,
) -> int:
    """
    Normal company users are always restricted to their own company.

    Global Admin may optionally supply a company ID when managing
    another company.
    """

    if current_user.is_global_admin:
        if requested_company_id is None:
            raise HTTPException(
                status_code=400,
                detail="company_id is required for Global Admin access",
            )

        return requested_company_id

    if current_user.company_id is None:
        raise HTTPException(
            status_code=403,
            detail="Your account is not assigned to a company",
        )

    return current_user.company_id


def require_farm_access(
    db: Session,
    current_user: models.AppUser,
    farm_id: int,
) -> BroilerFarm:
    farm = (
        db.query(BroilerFarm)
        .filter(BroilerFarm.id == farm_id)
        .first()
    )

    if not farm:
        raise HTTPException(
            status_code=404,
            detail="Broiler farm not found",
        )

    if current_user.is_global_admin:
        return farm

    if farm.company_id != current_user.company_id:
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if current_user.is_company_admin:
        return farm

    access = (
        db.query(models.UserFarmAccess)
        .filter(
            models.UserFarmAccess.user_id == current_user.id,
            models.UserFarmAccess.farm_id == farm_id,
        )
        .first()
    )

    if not access:
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this farm",
        )

    return farm

def build_shed_response(
    shed: BroilerShed,
    farm_name: str | None = None,
) -> BroilerShedOut:
    """
    Return every current Shed model field.

    This prevents newly added setup fields from disappearing when
    the frontend reloads the Shed Register after saving.
    """
    data = {
        column.name: getattr(shed, column.name)
        for column in shed.__table__.columns
    }

    data["farm_name"] = farm_name

    return BroilerShedOut.model_validate(data)

def build_daily_performance_response(
    entry: BroilerDailyPerformance,
    cumulative_mortality_birds: int | None = None,
):
    plan = entry.placement_plan
    farm = plan.farm if plan else None
    shed = plan.shed if plan else None

    opening_birds = entry.opening_birds or 0
    mortality_birds = entry.mortality_birds or 0
    cull_birds = entry.cull_birds or 0
    closing_birds = entry.closing_birds or 0
    feed_kg = float(entry.feed_kg) if entry.feed_kg is not None else 0

    daily_mortality_pct = None
    if opening_birds > 0:
        daily_mortality_pct = (mortality_birds / opening_birds) * 100

    cumulative_mortality_pct = None
    if opening_birds > 0 and cumulative_mortality_birds is not None:
        cumulative_mortality_pct = (cumulative_mortality_birds / opening_birds) * 100

    feed_per_bird_g = None
    if closing_birds > 0 and feed_kg > 0:
        feed_per_bird_g = (feed_kg * 1000) / closing_birds

    return BroilerDailyPerformanceOut(
        id=entry.id,
        company_id=entry.company_id,
        placement_plan_id=entry.placement_plan_id,

        farm_name=farm.farm_name if farm else None,
        shed_name=shed.shed_name if shed else None,
        cycle_code=plan.cycle_code if plan else None,

        entry_date=entry.entry_date,
        age_days=entry.age_days,

        opening_birds=entry.opening_birds,

        mortality_front=entry.mortality_front or 0,
        mortality_middle=entry.mortality_middle or 0,
        mortality_back=entry.mortality_back or 0,
        mortality_other=entry.mortality_other or 0,
        mortality_birds=entry.mortality_birds or 0,

        cull_legs=entry.cull_legs or 0,
        cull_runts=entry.cull_runts or 0,
        cull_beak=entry.cull_beak or 0,
        cull_other=entry.cull_other or 0,
        cull_birds=entry.cull_birds or 0,

        closing_birds=entry.closing_birds,

        feed_kg=float(entry.feed_kg) if entry.feed_kg is not None else None,
        water_litres=float(entry.water_litres) if entry.water_litres is not None else None,
        avg_weight_kg=float(entry.avg_weight_kg) if entry.avg_weight_kg is not None else None,
        body_weight_kg=float(entry.avg_weight_kg) if entry.avg_weight_kg is not None else None,

        daily_mortality_pct=daily_mortality_pct,
        cumulative_mortality_birds=cumulative_mortality_birds,
        cumulative_mortality_pct=cumulative_mortality_pct,
        feed_per_bird_g=feed_per_bird_g,

        notes=entry.notes,
        last_saved_by=entry.last_saved_by,
        last_saved_at=entry.last_saved_at,
    )



def ensure_module_access_schema() -> None:
    """
    Add the farm classification column to existing databases.

    Base.metadata.create_all() creates the new user_module_access table,
    but it does not add columns to an existing broiler_farms table.
    """
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "broiler_farms" not in table_names:
        return

    farm_columns = {
        column["name"]
        for column in inspector.get_columns("broiler_farms")
    }

    if "farm_type" not in farm_columns:
        with engine.begin() as connection:
            connection.execute(
                text(
                    "ALTER TABLE broiler_farms "
                    "ADD COLUMN farm_type VARCHAR(50) "
                    "NOT NULL DEFAULT 'broiler'"
                )
            )

    # Classify existing demonstration/master-data farms using their names.
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE broiler_farms
                SET farm_type = CASE
                    WHEN LOWER(farm_name) LIKE '%breeder rearing%'
                        THEN 'breeder_rearing'
                    WHEN LOWER(farm_name) LIKE '%breeder layer%'
                        THEN 'breeder_layers'
                    WHEN LOWER(farm_name) LIKE '%rearing%'
                        THEN 'layer_rearing'
                    WHEN LOWER(farm_name) LIKE '%layer%'
                        THEN 'commercial_layers'
                    ELSE COALESCE(NULLIF(farm_type, ''), 'broiler')
                END
                """
            )
        )



def ensure_broiler_shed_setup_schema() -> None:
    """
    Bring older broiler_sheds tables up to the current Shed setup model.

    SQLAlchemy's create_all() creates missing tables but does not add new
    columns to tables that already exist. Older Render/PostgreSQL databases
    therefore need these setup fields added explicitly before any ORM query
    loads BroilerShed.
    """
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "broiler_sheds" not in table_names:
        return

    existing_columns = {
        column["name"]
        for column in inspector.get_columns("broiler_sheds")
    }

    # Keep these definitions portable across PostgreSQL and SQLite.
    # shed_type mirrors the current model's required default; all other
    # expanded setup fields are nullable.
    column_sql = {
        "shed_type": "VARCHAR(80) NOT NULL DEFAULT 'Broiler'",
        "housing_system": "VARCHAR(80)",
        "capacity_birds": "INTEGER",
        "length_m": "NUMERIC(10, 2)",
        "width_m": "NUMERIC(10, 2)",
        "number_of_levels": "INTEGER",
        "number_of_sections": "INTEGER",
        "ventilation_type": "TEXT",
        "cooling_system": "TEXT",
        "heating_system": "TEXT",
        "lighting_system": "TEXT",
        "water_system": "TEXT",
        "feeder_system": "TEXT",
        "nest_type": "TEXT",
        "egg_collection_system": "TEXT",
        "manure_system": "TEXT",
        "year_commissioned": "INTEGER",
        "male_female_support": "VARCHAR(80)",
        "environmental_controller": "TEXT",
        "controller_id": "TEXT",
        "water_meter_id": "TEXT",
        "power_meter_id": "TEXT",
        "notes": "TEXT",
    }

    with engine.begin() as connection:
        for column_name, sql_type in column_sql.items():
            if column_name not in existing_columns:
                connection.execute(
                    text(
                        f"ALTER TABLE broiler_sheds "
                        f"ADD COLUMN {column_name} {sql_type}"
                    )
                )


def ensure_commercial_layer_transfer_schema() -> None:
    """
    Upgrade an existing commercial_layer_flocks table created by the
    performance-graph foundation.

    SQLAlchemy create_all() creates new tables but does not add columns to
    an existing table, so the source rearing link must be added explicitly.
    The column remains nullable for any historical/manual commercial-layer
    rows that predate the transfer workflow.
    """
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "commercial_layer_flocks" not in table_names:
        return

    columns = {
        column["name"]
        for column in inspector.get_columns("commercial_layer_flocks")
    }

    with engine.begin() as connection:
        if "source_rearing_flock_id" not in columns:
            connection.execute(
                text(
                    "ALTER TABLE commercial_layer_flocks "
                    "ADD COLUMN source_rearing_flock_id INTEGER"
                )
            )

        # PostgreSQL supports a partial unique index, which protects
        # transferred flocks while still allowing older rows with NULL.
        if engine.dialect.name == "postgresql":
            connection.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS "
                    "uq_commercial_layer_source_rearing_flock "
                    "ON commercial_layer_flocks "
                    "(source_rearing_flock_id) "
                    "WHERE source_rearing_flock_id IS NOT NULL"
                )
            )
        elif engine.dialect.name == "sqlite":
            connection.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS "
                    "uq_commercial_layer_source_rearing_flock "
                    "ON commercial_layer_flocks "
                    "(source_rearing_flock_id)"
                )
            )

def ensure_commercial_layer_operational_schema() -> None:
    """
    Add Commercial Layers flock-planning and egg-quality columns to
    existing databases.

    Base.metadata.create_all() creates new tables, but it does not add
    columns to existing tables.
    """
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    with engine.begin() as connection:
        if "commercial_layer_flocks" in table_names:
            flock_columns = {
                column["name"]
                for column in inspector.get_columns(
                    "commercial_layer_flocks"
                )
            }

            if "planned_depletion_date" not in flock_columns:
                connection.execute(
                    text(
                        "ALTER TABLE commercial_layer_flocks "
                        "ADD COLUMN planned_depletion_date DATE"
                    )
                )

        if "commercial_layer_daily_performance" in table_names:
            performance_columns = {
                column["name"]
                for column in inspector.get_columns(
                    "commercial_layer_daily_performance"
                )
            }

            column_sql = {
                "saleable_eggs": (
                    "INTEGER NOT NULL DEFAULT 0"
                ),
                "seconds": (
                    "INTEGER NOT NULL DEFAULT 0"
                ),
                "cracks": (
                    "INTEGER NOT NULL DEFAULT 0"
                ),
                "rejects": (
                    "INTEGER NOT NULL DEFAULT 0"
                ),
            }

            for column_name, sql_type in column_sql.items():
                if column_name not in performance_columns:
                    connection.execute(
                        text(
                            "ALTER TABLE "
                            "commercial_layer_daily_performance "
                            f"ADD COLUMN {column_name} {sql_type}"
                        )
                    )

def ensure_layer_transfer_schema() -> None:
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "layer_rearing_flocks" not in table_names:
        return

    columns = {
        column["name"]
        for column in inspector.get_columns("layer_rearing_flocks")
    }

    column_sql = {
        "actual_transfer_date": "DATE",
        "birds_transferred": "INTEGER",
        "transfer_notes": "TEXT",
        "transferred_by": "VARCHAR(255)",
        "transferred_at": "TIMESTAMP",
    }

    with engine.begin() as connection:
        for column_name, sql_type in column_sql.items():
            if column_name not in columns:
                connection.execute(
                    text(
                        f"ALTER TABLE layer_rearing_flocks "
                        f"ADD COLUMN {column_name} {sql_type}"
                    )
                )


def ensure_breeder_transfer_schema() -> None:
    """
    Add transfer audit columns to existing breeder rearing databases.

    Base.metadata.create_all() creates the new breeder_production_flocks
    table, but does not add columns to an existing breeder_rearing_flocks
    table.
    """
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "breeder_rearing_flocks" not in table_names:
        return

    columns = {
        column["name"]
        for column in inspector.get_columns("breeder_rearing_flocks")
    }

    column_sql = {
        "actual_transfer_date": "DATE",
        "females_transferred": "INTEGER",
        "males_transferred": "INTEGER",
        "transfer_notes": "TEXT",
        "transferred_by": "VARCHAR(255)",
        "transferred_at": "TIMESTAMP",
    }

    with engine.begin() as connection:
        for column_name, sql_type in column_sql.items():
            if column_name not in columns:
                connection.execute(
                    text(
                        f"ALTER TABLE breeder_rearing_flocks "
                        f"ADD COLUMN {column_name} {sql_type}"
                    )
                )


def ensure_user_activity_schema() -> None:
    """Add user-usage summary columns to existing app_users tables."""
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    if "app_users" not in table_names:
        return

    existing_columns = {
        column["name"]
        for column in inspector.get_columns("app_users")
    }

    column_sql = {
        "last_active_at": "TIMESTAMP",
        "last_module": "VARCHAR(80)",
        "last_page": "VARCHAR(255)",
    }

    with engine.begin() as connection:
        for column_name, sql_type in column_sql.items():
            if column_name not in existing_columns:
                connection.execute(
                    text(
                        f"ALTER TABLE app_users "
                        f"ADD COLUMN {column_name} {sql_type}"
                    )
                )


def repair_shed_company_links(db: Session) -> int:
    """
    Align every shed's company_id with its parent farm.

    Older sandbox data could contain sheds created before the
    farm/company validation was enforced. This repair is safe because
    the farm is the authoritative owner of the shed.
    """
    mismatched = (
        db.query(BroilerShed)
        .join(
            BroilerFarm,
            BroilerFarm.id == BroilerShed.farm_id,
        )
        .filter(
            BroilerShed.company_id
            != BroilerFarm.company_id
        )
        .all()
    )

    repaired = 0

    for shed in mismatched:
        farm = (
            db.query(BroilerFarm)
            .filter(BroilerFarm.id == shed.farm_id)
            .first()
        )

        if not farm:
            continue

        shed.company_id = farm.company_id
        repaired += 1

    if repaired:
        db.commit()

    return repaired

@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    ensure_module_access_schema()
    ensure_broiler_shed_setup_schema()
    ensure_commercial_layer_transfer_schema()
    ensure_commercial_layer_operational_schema()
    ensure_layer_transfer_schema()
    ensure_breeder_transfer_schema()
    ensure_user_activity_schema()

    should_seed_demo_data = (
        os.getenv("SEED_DEMO_DATA", "false").strip().lower()
        in {"1", "true", "yes", "on"}
    )

    db = SessionLocal()

    try:
        repair_shed_company_links(db)

        if should_seed_demo_data:
            seed_demo_data(db)
    finally:
        db.close()


@app.get("/api/health")
def health():
    return {"status": "ok", "module": "broilers"}

@app.get(
    "/api/broilers/demand-plans",
    response_model=list[BroilerDemandPlanOut],
)
def list_demand_plans(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.company_id
            == resolved_company_id
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerPlacementPlan.farm_id.in_(
                permitted_farm_ids
            )
        )

    plans = (
        query
        .order_by(
            BroilerPlacementPlan.placement_date.asc(),
            BroilerPlacementPlan.id.asc(),
        )
        .all()
    )

    return [
        build_plan_response(plan)
        for plan in plans
    ]


@app.patch(
    "/api/broilers/demand-plans/{plan_id}",
    response_model=BroilerDemandPlanOut,
)
def update_demand_plan(
    plan_id: int,
    payload: dict = Body(...),
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id == plan_id
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail="Broiler demand plan not found",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    if (
        not current_user.is_global_admin
        and plan.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    # Validate the standard demand-plan fields while retaining farm_id and
    # shed_id, which older BroilerDemandPlanPatch schemas may ignore.
    validated_payload = BroilerDemandPlanPatch.model_validate(payload)
    data = validated_payload.model_dump(exclude_unset=True)

    requested_farm_id = payload.get("farm_id", plan.farm_id)
    requested_shed_id = payload.get("shed_id", plan.shed_id)

    try:
        requested_farm_id = int(requested_farm_id)
        requested_shed_id = int(requested_shed_id)
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail="farm_id and shed_id must be valid integers",
        )

    target_farm = require_farm_access(
        db,
        current_user,
        requested_farm_id,
    )

    target_shed = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.id == requested_shed_id,
            BroilerShed.active == True,
        )
        .first()
    )

    if not target_shed:
        raise HTTPException(
            status_code=404,
            detail="Broiler shed not found or inactive",
        )

    if (
        target_shed.farm_id != target_farm.id
        or target_shed.company_id != target_farm.company_id
        or target_farm.company_id != plan.company_id
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "The selected shed does not belong to the selected farm "
                "and company."
            ),
        )

    plan.farm_id = target_farm.id
    plan.shed_id = target_shed.id

    for field, value in data.items():
        if field not in {"farm_id", "shed_id", "company_id"}:
            setattr(plan, field, value)

    plan.last_saved_by = current_user.full_name
    plan.last_saved_at = datetime.utcnow()

    db.commit()
    db.refresh(plan)

    return build_plan_response(plan)


@app.post(
    "/api/broilers/demand-plans",
    response_model=BroilerDemandPlanOut,
)
def create_demand_plan(
    payload: BroilerDemandPlanCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    farm = require_farm_access(
        db,
        current_user,
        payload.farm_id,
    )

    shed = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.id == payload.shed_id
        )
        .first()
    )

    if not shed:
        raise HTTPException(
            status_code=404,
            detail="Broiler shed not found",
        )

    if (
        shed.farm_id != farm.id
        or shed.company_id != farm.company_id
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "The selected shed does not belong "
                "to the selected farm."
            ),
        )

    plan = BroilerPlacementPlan(
        company_id=farm.company_id,
        farm_id=farm.id,
        shed_id=shed.id,
        cycle_code=payload.cycle_code,
        placement_date=payload.placement_date,
        planned_birds=payload.planned_birds,
        target_density_kg_m2=(
            payload.target_density_kg_m2
        ),
        target_lw_kg=payload.target_lw_kg,
        growout_days=payload.growout_days,
        chick_allowance_pct=(
            payload.chick_allowance_pct
        ),
        notes=payload.notes,
        status=payload.status,
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(plan)
    db.commit()
    db.refresh(plan)

    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id == plan.id,
            BroilerPlacementPlan.company_id
            == farm.company_id,
        )
        .first()
    )

    return build_plan_response(plan)


@app.delete("/api/broilers/demand-plans/{plan_id}")
def delete_demand_plan(
    plan_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    plan = (
        db.query(BroilerPlacementPlan)
        .filter(
            BroilerPlacementPlan.id == plan_id
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail="Broiler demand plan not found",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    linked_entries = (
        db.query(BroilerDailyPerformance)
        .filter(
            BroilerDailyPerformance.company_id
            == plan.company_id,
            BroilerDailyPerformance.placement_plan_id
            == plan.id,
        )
        .count()
    )

    if linked_entries > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot delete this plan because it has "
                "daily performance records."
            ),
        )

    db.delete(plan)
    db.commit()

    return {
        "deleted": True,
        "id": plan_id,
    }


@app.post(
    "/api/broilers/demand-plans/new-row",
    response_model=BroilerDemandPlanOut,
)
def create_broiler_demand_plan_new_row(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerShed)
        .join(
            BroilerFarm,
            BroilerFarm.id == BroilerShed.farm_id,
        )
        .filter(
            BroilerShed.active == True,
            BroilerShed.company_id
            == resolved_company_id,
            BroilerFarm.company_id
            == resolved_company_id,
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerShed.farm_id.in_(
                permitted_farm_ids
            )
        )

    shed = (
        query
        .order_by(BroilerShed.id.asc())
        .first()
    )

    if not shed:
        raise HTTPException(
            status_code=400,
            detail=(
                "No active broiler sheds are available "
                "for this user and company."
            ),
        )

    farm = require_farm_access(
        db,
        current_user,
        shed.farm_id,
    )

    existing_count = (
        db.query(BroilerPlacementPlan)
        .filter(
            BroilerPlacementPlan.company_id
            == resolved_company_id
        )
        .count()
    )

    next_number = existing_count + 1

    plan = BroilerPlacementPlan(
        company_id=resolved_company_id,
        farm_id=farm.id,
        shed_id=shed.id,
        cycle_code=f"BR-NEW-{next_number:03d}",
        placement_date=None,
        planned_birds=None,
        target_density_kg_m2=(
            shed.default_density_kg_m2
        ),
        target_lw_kg=shed.default_target_lw_kg,
        growout_days=shed.default_growout_days,
        chick_allowance_pct=1.5,
        notes="",
        status="Draft",
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(plan)
    db.commit()
    db.refresh(plan)

    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id == plan.id,
            BroilerPlacementPlan.company_id
            == resolved_company_id,
        )
        .first()
    )

    return build_plan_response(plan)

@app.get(
    "/api/broilers/farms",
    response_model=list[BroilerFarmOut],
)
def list_broiler_farms(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerFarm)
        .filter(
            BroilerFarm.company_id == resolved_company_id
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerFarm.id.in_(permitted_farm_ids)
        )

    return (
        query
        .order_by(BroilerFarm.farm_name.asc())
        .all()
    )


@app.post(
    "/api/broilers/farms",
    response_model=BroilerFarmOut,
)
def create_broiler_farm(
    payload: BroilerFarmCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    resolved_company_id = resolve_company_id(
        current_user,
        payload.company_id,
    )

    if not current_user.is_global_admin:
        resolved_company_id = current_user.company_id

    company = (
        db.query(Company)
        .filter(
            Company.id == resolved_company_id,
            Company.active == True,
        )
        .first()
    )

    if not company:
        raise HTTPException(
            status_code=404,
            detail="Company not found or inactive",
        )

    existing = (
        db.query(BroilerFarm)
        .filter(
            BroilerFarm.company_id
            == resolved_company_id,
            BroilerFarm.farm_name
            == payload.farm_name.strip(),
        )
        .first()
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail=(
                "A farm with this name already exists "
                "for the company."
            ),
        )

    allowed_farm_types = {
        "broiler",
        "breeder_rearing",
        "breeder_layers",
        "layer_rearing",
        "commercial_layers",
        "hatchery",
        "feed_mill",
        "grading",
        "processing",
    }

    farm_type = payload.farm_type.strip().lower()

    if farm_type not in allowed_farm_types:
        raise HTTPException(
            status_code=400,
            detail="Invalid farm type",
        )

    farm = BroilerFarm(
        company_id=resolved_company_id,
        farm_name=payload.farm_name.strip(),
        farm_type=farm_type,
        farm_code=(
            payload.farm_code.strip()
            if payload.farm_code
            else None
        ),
        active=payload.active,
    )

    db.add(farm)
    db.commit()
    db.refresh(farm)

    return farm


@app.patch(
    "/api/broilers/farms/{farm_id}",
    response_model=BroilerFarmOut,
)
def update_broiler_farm(
    farm_id: int,
    payload: BroilerFarmPatch,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    farm = require_farm_access(
        db,
        current_user,
        farm_id,
    )

    data = payload.model_dump(exclude_unset=True)

    if "farm_name" in data and data["farm_name"]:
        farm_name = data["farm_name"].strip()

        duplicate = (
            db.query(BroilerFarm)
            .filter(
                BroilerFarm.company_id
                == farm.company_id,
                BroilerFarm.farm_name
                == farm_name,
                BroilerFarm.id != farm.id,
            )
            .first()
        )

        if duplicate:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Another farm with this name already "
                    "exists for the company."
                ),
            )

        data["farm_name"] = farm_name

    if "farm_code" in data and data["farm_code"]:
        data["farm_code"] = data["farm_code"].strip()

    if "farm_type" in data and data["farm_type"]:
        allowed_farm_types = {
            "broiler",
            "breeder_rearing",
            "breeder_layers",
            "layer_rearing",
            "commercial_layers",
            "hatchery",
            "feed_mill",
            "grading",
            "processing",
        }

        farm_type = data["farm_type"].strip().lower()

        if farm_type not in allowed_farm_types:
            raise HTTPException(
                status_code=400,
                detail="Invalid farm type",
            )

        data["farm_type"] = farm_type

    for field, value in data.items():
        setattr(farm, field, value)

    db.commit()
    db.refresh(farm)

    return farm


@app.delete("/api/broilers/farms/{farm_id}")
def delete_broiler_farm(
    farm_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_global_admin:
        raise HTTPException(
            status_code=403,
            detail="Global Admin access required",
        )

    farm = require_farm_access(
        db,
        current_user,
        farm_id,
    )

    linked_sheds = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.farm_id == farm.id,
            BroilerShed.company_id == farm.company_id,
        )
        .count()
    )

    if linked_sheds > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot delete farm because it has linked "
                "sheds. Set the farm inactive instead."
            ),
        )

    db.delete(farm)
    db.commit()

    return {
        "deleted": True,
        "id": farm_id,
    }

@app.get(
    "/api/broilers/sheds",
    response_model=list[BroilerShedOut],
)
def list_broiler_sheds(
    company_id: int | None = None,
    farm_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(
            BroilerShed,
            BroilerFarm.farm_name,
        )
        .join(
            BroilerFarm,
            BroilerFarm.id == BroilerShed.farm_id,
        )
        .filter(
            BroilerFarm.company_id
            == resolved_company_id,
        )
    )

    if farm_id is not None:
        require_farm_access(
            db,
            current_user,
            farm_id,
        )

        query = query.filter(
            BroilerShed.farm_id == farm_id
        )

    elif not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerShed.farm_id.in_(permitted_farm_ids)
        )

    sheds = (
        query
        .order_by(
            BroilerFarm.farm_name.asc(),
            BroilerShed.shed_name.asc(),
        )
        .all()
    )

    return [
        build_shed_response(
            shed,
            farm_name=farm_name,
        )
        for shed, farm_name in sheds
    ]


@app.post(
    "/api/broilers/sheds",
    response_model=BroilerShedOut,
)
def create_broiler_shed(
    payload: BroilerShedCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    farm = require_farm_access(
        db,
        current_user,
        payload.farm_id,
    )

    if (
        current_user.is_global_admin
        and payload.company_id != farm.company_id
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "The selected company does not match "
                "the farm company."
            ),
        )

    existing = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.company_id == farm.company_id,
            BroilerShed.farm_id == farm.id,
            BroilerShed.shed_name
            == payload.shed_name.strip(),
        )
        .first()
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail=(
                "A shed with this name already exists "
                "on the selected farm."
            ),
        )

    shed = BroilerShed(
        company_id=farm.company_id,
        farm_id=farm.id,
        shed_name=payload.shed_name.strip(),
        shed_code=(
            payload.shed_code.strip()
            if payload.shed_code
            else None
        ),
        floor_area_m2=payload.floor_area_m2,
        default_density_kg_m2=(
            payload.default_density_kg_m2
        ),
        default_target_lw_kg=(
            payload.default_target_lw_kg
        ),
        default_growout_days=(
            payload.default_growout_days
        ),
        active=payload.active,
    )

    db.add(shed)
    db.commit()
    db.refresh(shed)

    return build_shed_response(
        shed,
        farm_name=farm.farm_name,
    )


@app.patch(
    "/api/broilers/sheds/{shed_id}",
    response_model=BroilerShedOut,
)
def update_broiler_shed(
    shed_id: int,
    payload: BroilerShedPatch,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    shed = (
        db.query(BroilerShed)
        .filter(BroilerShed.id == shed_id)
        .first()
    )

    if not shed:
        raise HTTPException(
            status_code=404,
            detail="Broiler shed not found",
        )

    require_farm_access(
        db,
        current_user,
        shed.farm_id,
    )

    data = payload.model_dump(exclude_unset=True)

    if "farm_id" in data:
        target_farm = require_farm_access(
            db,
            current_user,
            data["farm_id"],
        )

        if target_farm.company_id != shed.company_id:
            raise HTTPException(
                status_code=400,
                detail=(
                    "A shed cannot be moved to a farm "
                    "belonging to another company."
                ),
            )

    target_farm_id = data.get(
        "farm_id",
        shed.farm_id,
    )

    if "shed_name" in data and data["shed_name"]:
        shed_name = data["shed_name"].strip()

        duplicate = (
            db.query(BroilerShed)
            .filter(
                BroilerShed.company_id
                == shed.company_id,
                BroilerShed.farm_id
                == target_farm_id,
                BroilerShed.shed_name
                == shed_name,
                BroilerShed.id != shed.id,
            )
            .first()
        )

        if duplicate:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Another shed with this name already "
                    "exists on the selected farm."
                ),
            )

        data["shed_name"] = shed_name

    if "shed_code" in data and data["shed_code"]:
        data["shed_code"] = data["shed_code"].strip()

    for field, value in data.items():
        setattr(shed, field, value)

    db.commit()
    db.refresh(shed)

    farm = require_farm_access(
        db,
        current_user,
        shed.farm_id,
    )

    return build_shed_response(
        shed,
        farm_name=farm.farm_name,
    )


@app.delete("/api/broilers/sheds/{shed_id}")
def delete_broiler_shed(
    shed_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_global_admin:
        raise HTTPException(
            status_code=403,
            detail="Global Admin access required",
        )

    shed = (
        db.query(BroilerShed)
        .filter(BroilerShed.id == shed_id)
        .first()
    )

    if not shed:
        raise HTTPException(
            status_code=404,
            detail="Broiler shed not found",
        )

    require_farm_access(
        db,
        current_user,
        shed.farm_id,
    )

    linked_plans = (
        db.query(BroilerPlacementPlan)
        .filter(
            BroilerPlacementPlan.shed_id == shed.id,
            BroilerPlacementPlan.company_id
            == shed.company_id,
        )
        .count()
    )

    if linked_plans > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot delete shed because it has linked "
                "placement plans. Set the shed inactive instead."
            ),
        )

    db.delete(shed)
    db.commit()

    return {
        "deleted": True,
        "id": shed_id,
    }

def recalculate_daily_performance_entry(entry: BroilerDailyPerformance):
    mortality_total = (
        int(entry.mortality_front or 0)
        + int(entry.mortality_middle or 0)
        + int(entry.mortality_back or 0)
        + int(entry.mortality_other or 0)
    )

    cull_total = (
        int(entry.cull_legs or 0)
        + int(entry.cull_runts or 0)
        + int(entry.cull_beak or 0)
        + int(entry.cull_other or 0)
    )

    entry.mortality_birds = mortality_total
    entry.cull_birds = cull_total

    if entry.opening_birds is not None:
        entry.closing_birds = (
            int(entry.opening_birds or 0)
            - mortality_total
            - cull_total
        )
    else:
        entry.closing_birds = None

    return entry
    
@app.get(
    "/api/broilers/performance",
    response_model=list[BroilerDailyPerformanceOut],
)
def list_broiler_performance(
    company_id: int | None = None,
    placement_plan_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerDailyPerformance)
        .options(
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.farm),
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.shed),
        )
        .join(
            BroilerPlacementPlan,
            BroilerPlacementPlan.id
            == BroilerDailyPerformance.placement_plan_id,
        )
        .filter(
            BroilerDailyPerformance.company_id
            == resolved_company_id,
            BroilerPlacementPlan.company_id
            == resolved_company_id,
        )
    )

    if placement_plan_id is not None:
        plan = (
            db.query(BroilerPlacementPlan)
            .filter(
                BroilerPlacementPlan.id
                == placement_plan_id,
                BroilerPlacementPlan.company_id
                == resolved_company_id,
            )
            .first()
        )

        if not plan:
            raise HTTPException(
                status_code=404,
                detail="Broiler placement plan not found",
            )

        require_farm_access(
            db,
            current_user,
            plan.farm_id,
        )

        query = query.filter(
            BroilerDailyPerformance.placement_plan_id
            == placement_plan_id
        )

    elif not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerPlacementPlan.farm_id.in_(
                permitted_farm_ids
            )
        )

    entries = (
        query
        .order_by(
            BroilerDailyPerformance.placement_plan_id.asc(),
            BroilerDailyPerformance.entry_date.asc(),
            BroilerDailyPerformance.id.asc(),
        )
        .all()
    )

    cumulative_by_plan: dict[int, int] = {}
    output = []

    for entry in entries:
        plan_id = entry.placement_plan_id

        cumulative_by_plan.setdefault(
            plan_id,
            0,
        )

        cumulative_by_plan[plan_id] += int(
            entry.mortality_birds or 0
        )

        output.append(
            build_daily_performance_response(
                entry,
                cumulative_mortality_birds=(
                    cumulative_by_plan[plan_id]
                ),
            )
        )

    return output


@app.post(
    "/api/broilers/performance",
    response_model=BroilerDailyPerformanceOut,
)
def create_broiler_performance(
    payload: BroilerDailyPerformanceCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id
            == payload.placement_plan_id
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail="Broiler placement plan not found",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    if (
        not current_user.is_global_admin
        and plan.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    existing = (
        db.query(BroilerDailyPerformance)
        .filter(
            BroilerDailyPerformance.company_id
            == plan.company_id,
            BroilerDailyPerformance.placement_plan_id
            == plan.id,
            BroilerDailyPerformance.entry_date
            == payload.entry_date,
        )
        .first()
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail=(
                "A performance entry already exists "
                "for this cycle and date."
            ),
        )

    data = payload.model_dump()

    data["company_id"] = plan.company_id

    if "body_weight_kg" in data:
        data["avg_weight_kg"] = data.pop(
            "body_weight_kg"
        )

    entry = BroilerDailyPerformance(**data)

    recalculate_daily_performance_entry(entry)

    entry.last_saved_by = current_user.full_name
    entry.last_saved_at = datetime.utcnow()

    db.add(entry)
    db.commit()
    db.refresh(entry)

    entry = (
        db.query(BroilerDailyPerformance)
        .options(
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.farm),
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerDailyPerformance.id == entry.id,
            BroilerDailyPerformance.company_id
            == plan.company_id,
        )
        .first()
    )

    return build_daily_performance_response(
        entry,
        cumulative_mortality_birds=(
            entry.mortality_birds or 0
        ),
    )


@app.patch(
    "/api/broilers/performance/{entry_id}",
    response_model=BroilerDailyPerformanceOut,
)
def update_broiler_performance(
    entry_id: int,
    payload: BroilerDailyPerformancePatch,
    expected_last_saved_at: str | None = Header(
        default=None,
        alias="X-OviCore-Expected-Last-Saved-At",
    ),
    mobile_sync: str | None = Header(
        default=None,
        alias="X-OviCore-Mobile-Sync",
    ),
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    entry = (
        db.query(BroilerDailyPerformance)
        .options(
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.farm),
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerDailyPerformance.id == entry_id
        )
        .first()
    )

    if not entry:
        raise HTTPException(
            status_code=404,
            detail="Broiler performance entry not found",
        )

    plan = entry.placement_plan

    if not plan:
        raise HTTPException(
            status_code=400,
            detail="Performance entry has no valid placement plan",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    if (
        not current_user.is_global_admin
        and entry.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if expected_last_saved_at is not None:
        actual_last_saved_at = (
            entry.last_saved_at.isoformat()
            if entry.last_saved_at is not None
            else ""
        )

        expected_normalised = (
            expected_last_saved_at
            .strip()
            .replace("Z", "+00:00")
        )

        actual_normalised = (
            actual_last_saved_at
            .strip()
            .replace("Z", "+00:00")
        )

        try:
            expected_dt = datetime.fromisoformat(
                expected_normalised
            )
            actual_dt = datetime.fromisoformat(
                actual_normalised
            )
            same_version = (
                expected_dt.replace(tzinfo=None)
                == actual_dt.replace(tzinfo=None)
            )
        except ValueError:
            same_version = (
                expected_normalised
                == actual_normalised
            )

        if not same_version:
            raise HTTPException(
                status_code=409,
                detail=(
                    "This performance entry changed in "
                    "OviCore after it was loaded on mobile. "
                    "Review the sync conflict before updating."
                ),
            )

    data = payload.model_dump(
        exclude_unset=True
    )

    if (
        mobile_sync is not None
        and mobile_sync.strip().lower()
        in {"1", "true", "yes", "on"}
    ):
        data = {
            field: value
            for field, value in data.items()
            if value is not None
        }

    if "body_weight_kg" in data:
        data["avg_weight_kg"] = data.pop(
            "body_weight_kg"
        )

    if (
        mobile_sync is not None
        and mobile_sync.strip().lower()
        in {"1", "true", "yes", "on"}
    ):
        protected_fields = []

        for field, incoming_value in data.items():
            if field in {
                "placement_plan_id",
                "entry_date",
                "age_days",
                "mortality_birds",
                "cull_birds",
                "closing_birds",
                "last_saved_by",
            }:
                continue

            existing_value = getattr(
                entry,
                field,
                None,
            )

            if (
                existing_value is None
                or incoming_value is None
            ):
                continue

            try:
                values_are_equal = (
                    float(existing_value)
                    == float(incoming_value)
                )
            except (TypeError, ValueError):
                values_are_equal = (
                    str(existing_value).strip()
                    == str(incoming_value).strip()
                )

            if not values_are_equal:
                protected_fields.append(field)

        if protected_fields:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Mobile sync cannot overwrite existing "
                    "OviCore values. Protected fields: "
                    + ", ".join(protected_fields)
                ),
            )

    if "entry_date" in data:
        duplicate = (
            db.query(BroilerDailyPerformance)
            .filter(
                BroilerDailyPerformance.company_id
                == entry.company_id,
                BroilerDailyPerformance.placement_plan_id
                == entry.placement_plan_id,
                BroilerDailyPerformance.entry_date
                == data["entry_date"],
                BroilerDailyPerformance.id != entry.id,
            )
            .first()
        )

        if duplicate:
            raise HTTPException(
                status_code=400,
                detail=(
                    "A performance entry already exists "
                    "for this cycle and date."
                ),
            )

    for field, value in data.items():
        setattr(entry, field, value)

    recalculate_daily_performance_entry(entry)

    entry.last_saved_by = current_user.full_name
    entry.last_saved_at = datetime.utcnow()

    db.commit()
    db.refresh(entry)

    cumulative_mortality = (
        db.query(BroilerDailyPerformance)
        .filter(
            BroilerDailyPerformance.company_id
            == entry.company_id,
            BroilerDailyPerformance.placement_plan_id
            == entry.placement_plan_id,
            BroilerDailyPerformance.entry_date
            <= entry.entry_date,
        )
        .with_entities(
            BroilerDailyPerformance.mortality_birds
        )
        .all()
    )

    cumulative_mortality_birds = sum(
        int(row[0] or 0)
        for row in cumulative_mortality
    )

    return build_daily_performance_response(
        entry,
        cumulative_mortality_birds=(
            cumulative_mortality_birds
        ),
    )


@app.delete("/api/broilers/performance/{entry_id}")
def delete_broiler_performance(
    entry_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    entry = (
        db.query(BroilerDailyPerformance)
        .options(
            joinedload(
                BroilerDailyPerformance.placement_plan
            )
        )
        .filter(
            BroilerDailyPerformance.id == entry_id
        )
        .first()
    )

    if not entry:
        raise HTTPException(
            status_code=404,
            detail="Broiler performance entry not found",
        )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    plan = entry.placement_plan

    if not plan:
        raise HTTPException(
            status_code=400,
            detail="Performance entry has no valid placement plan",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    db.delete(entry)
    db.commit()

    return {
        "deleted": True,
        "id": entry_id,
    }


@app.post(
    "/api/broilers/performance/recalculate-cycle/{placement_plan_id}"
)
def recalculate_broiler_performance_cycle(
    placement_plan_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    plan = (
        db.query(BroilerPlacementPlan)
        .filter(
            BroilerPlacementPlan.id
            == placement_plan_id
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail="Broiler placement plan not found",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    entries = (
        db.query(BroilerDailyPerformance)
        .filter(
            BroilerDailyPerformance.company_id
            == plan.company_id,
            BroilerDailyPerformance.placement_plan_id
            == placement_plan_id,
        )
        .order_by(
            BroilerDailyPerformance.entry_date.asc(),
            BroilerDailyPerformance.age_days.asc(),
            BroilerDailyPerformance.id.asc(),
        )
        .all()
    )

    if not entries:
        raise HTTPException(
            status_code=404,
            detail=(
                "No performance entries found "
                "for this cycle."
            ),
        )

    previous_closing_birds = None

    for index, entry in enumerate(entries):
        if (
            index > 0
            and previous_closing_birds is not None
        ):
            entry.opening_birds = (
                previous_closing_birds
            )

        recalculate_daily_performance_entry(entry)

        previous_closing_birds = entry.closing_birds
        entry.last_saved_at = datetime.utcnow()
        entry.last_saved_by = current_user.full_name

    db.commit()

    return {
        "ok": True,
        "placement_plan_id": placement_plan_id,
        "rows_recalculated": len(entries),
    }


# ---------------------------------------------------------------------
# Broiler Paper Capture
# Handwritten AM / PM shed sheet -> AI extraction -> human review -> save
# ---------------------------------------------------------------------

PAPER_CAPTURE_TEMPLATE_VERSION = "broiler-v1"
PAPER_CAPTURE_MODEL = os.getenv(
    "PAPER_CAPTURE_MODEL",
    "gpt-5.6-luna",
).strip()


def _paper_capture_template_parts(
    template_id: str,
) -> tuple[int, date] | None:
    match = re.fullmatch(
        r"BRS-(\d+)-(\d{8})",
        (template_id or "").strip().upper(),
    )

    if not match:
        return None

    try:
        plan_id = int(match.group(1))
        entry_date = datetime.strptime(
            match.group(2),
            "%Y%m%d",
        ).date()
    except (TypeError, ValueError):
        return None

    return plan_id, entry_date


def _paper_capture_number(value):
    if value is None or value == "":
        return None

    try:
        number = float(value)
    except (TypeError, ValueError):
        return None

    return number


def _paper_capture_int(value) -> int | None:
    number = _paper_capture_number(value)
    if number is None:
        return None
    return int(round(number))


def _paper_capture_float(value) -> float | None:
    number = _paper_capture_number(value)
    if number is None:
        return None
    return float(number)


def _paper_capture_sum_int(*values) -> int:
    return sum(
        value
        for value in (
            _paper_capture_int(item)
            for item in values
        )
        if value is not None
    )


def _paper_capture_sum_float(*values) -> float | None:
    cleaned = [
        value
        for value in (
            _paper_capture_float(item)
            for item in values
        )
        if value is not None
    ]

    if not cleaned:
        return None

    return round(sum(cleaned), 2)


def _paper_capture_proposed(
    source: dict,
) -> tuple[dict, list[str]]:
    warnings: list[str] = []

    opening_am = _paper_capture_int(
        source.get("opening_birds_am")
    )
    opening_pm = _paper_capture_int(
        source.get("opening_birds_pm")
    )

    opening_birds = (
        opening_am
        if opening_am is not None
        else opening_pm
    )

    if (
        opening_am is not None
        and opening_pm is not None
        and opening_am != opening_pm
    ):
        warnings.append(
            "AM and PM Opening Birds differ. "
            "OviCore proposed the AM value; confirm it during review."
        )

    body_am = _paper_capture_float(
        source.get("body_weight_kg_am")
    )
    body_pm = _paper_capture_float(
        source.get("body_weight_kg_pm")
    )

    body_weight = (
        body_pm
        if body_pm is not None
        else body_am
    )

    if (
        body_am is not None
        and body_pm is not None
        and abs(body_am - body_pm) > 0.02
    ):
        warnings.append(
            "Both AM and PM bodyweights were recorded and differ. "
            "OviCore proposed the PM value; confirm it during review."
        )

    observations = (
        str(source.get("observations") or "").strip()
    )
    actions = (
        str(source.get("actions_taken") or "").strip()
    )

    note_parts = []
    if observations:
        note_parts.append(
            f"Paper observations: {observations}"
        )
    if actions:
        note_parts.append(
            f"Actions taken: {actions}"
        )

    proposed = {
        "opening_birds": opening_birds,

        "mortality_front": _paper_capture_sum_int(
            source.get("mortality_front_am"),
            source.get("mortality_front_pm"),
        ),
        "mortality_middle": _paper_capture_sum_int(
            source.get("mortality_middle_am"),
            source.get("mortality_middle_pm"),
        ),
        "mortality_back": _paper_capture_sum_int(
            source.get("mortality_back_am"),
            source.get("mortality_back_pm"),
        ),
        "mortality_other": _paper_capture_sum_int(
            source.get("mortality_other_am"),
            source.get("mortality_other_pm"),
        ),

        "cull_legs": _paper_capture_sum_int(
            source.get("cull_legs_am"),
            source.get("cull_legs_pm"),
        ),
        "cull_runts": _paper_capture_sum_int(
            source.get("cull_runts_am"),
            source.get("cull_runts_pm"),
        ),
        "cull_beak": _paper_capture_sum_int(
            source.get("cull_beak_am"),
            source.get("cull_beak_pm"),
        ),
        "cull_other": _paper_capture_sum_int(
            source.get("cull_other_am"),
            source.get("cull_other_pm"),
        ),

        "feed_kg": _paper_capture_sum_float(
            source.get("feed_kg_am"),
            source.get("feed_kg_pm"),
        ),
        "water_litres": _paper_capture_sum_float(
            source.get("water_litres_am"),
            source.get("water_litres_pm"),
        ),
        "body_weight_kg": body_weight,
        "notes": "\n".join(note_parts) or None,
    }

    return proposed, warnings


def _paper_capture_json_schema() -> dict:
    nullable_integer = {
        "anyOf": [
            {"type": "integer"},
            {"type": "null"},
        ]
    }
    nullable_number = {
        "anyOf": [
            {"type": "number"},
            {"type": "null"},
        ]
    }
    nullable_string = {
        "anyOf": [
            {"type": "string"},
            {"type": "null"},
        ]
    }

    integer_fields = [
        "opening_birds_am",
        "opening_birds_pm",
        "mortality_front_am",
        "mortality_front_pm",
        "mortality_middle_am",
        "mortality_middle_pm",
        "mortality_back_am",
        "mortality_back_pm",
        "mortality_other_am",
        "mortality_other_pm",
        "cull_legs_am",
        "cull_legs_pm",
        "cull_runts_am",
        "cull_runts_pm",
        "cull_beak_am",
        "cull_beak_pm",
        "cull_other_am",
        "cull_other_pm",
    ]

    number_fields = [
        "feed_kg_am",
        "feed_kg_pm",
        "water_litres_am",
        "water_litres_pm",
        "body_weight_kg_am",
        "body_weight_kg_pm",
    ]

    confidence_fields = [
        "template_id",
        *integer_fields,
        *number_fields,
        "observations",
        "actions_taken",
    ]

    properties = {
        "template_id": {"type": "string"},
        **{
            field: nullable_integer
            for field in integer_fields
        },
        **{
            field: nullable_number
            for field in number_fields
        },
        "observations": nullable_string,
        "actions_taken": nullable_string,
        "confidence": {
            "type": "object",
            "properties": {
                field: {
                    "type": "number",
                    "minimum": 0,
                    "maximum": 1,
                }
                for field in confidence_fields
            },
            "required": confidence_fields,
            "additionalProperties": False,
        },
    }

    return {
        "type": "object",
        "properties": properties,
        "required": list(properties.keys()),
        "additionalProperties": False,
    }


def _paper_capture_extract_with_openai(
    image_bytes: bytes,
    content_type: str,
) -> dict:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()

    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=(
                "Paper Capture AI is not configured. "
                "Set OPENAI_API_KEY on the backend service."
            ),
        )

    encoded = base64.b64encode(
        image_bytes
    ).decode("ascii")

    data_url = (
        f"data:{content_type};base64,{encoded}"
    )

    prompt = """
You are reading an OviCore Broiler Shed Daily Record V1.
The form has printed labels and handwritten AM / PM entries.

Return only the requested structured data.

Rules:
- Read the printed Template ID in the top-right area. It looks like
  BRS-<placement_plan_id>-<YYYYMMDD>.
- Read handwritten values only from their labelled AM / PM cells.
- Do not invent missing numbers. Use null when a cell is blank, crossed
  out, unreadable, or genuinely uncertain.
- Whole bird counts must be integers.
- Feed and water may contain decimals.
- Bodyweight is kilograms.
- observations is the handwritten text under Observations / Issues Noticed.
- actions_taken is the handwritten text under Actions Taken Today.
- confidence values are 0 to 1 and must reflect confidence in the exact
  value read from that cell.
- Ignore the printed OK / Issue, Normal / Issue and No / Yes guide text
  for this V1 extraction.
"""

    request_body = {
        "model": PAPER_CAPTURE_MODEL,
        "input": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": prompt,
                    },
                    {
                        "type": "input_image",
                        "image_url": data_url,
                        "detail": "high",
                    },
                ],
            }
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "ovicore_broiler_paper_capture",
                "strict": True,
                "schema": _paper_capture_json_schema(),
            }
        },
    }

    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(request_body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(
            request,
            timeout=90,
        ) as response:
            response_json = json.loads(
                response.read().decode("utf-8")
            )
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode(
            "utf-8",
            errors="replace",
        )
        raise HTTPException(
            status_code=502,
            detail=(
                "Paper Capture AI request failed. "
                f"OpenAI returned {exc.code}: {detail[:600]}"
            ),
        )
    except Exception as exc:
        raise HTTPException(
            status_code=502,
            detail=f"Paper Capture AI request failed: {exc}",
        )

    output_text = ""

    for output_item in response_json.get(
        "output",
        [],
    ):
        for content_item in output_item.get(
            "content",
            [],
        ):
            if content_item.get("type") == "output_text":
                output_text = content_item.get(
                    "text",
                    "",
                )
                break
        if output_text:
            break

    if not output_text:
        raise HTTPException(
            status_code=502,
            detail="Paper Capture AI returned no readable extraction.",
        )

    try:
        return json.loads(output_text)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=502,
            detail=(
                "Paper Capture AI returned invalid structured data: "
                f"{exc}"
            ),
        )


def _paper_capture_overall_confidence(
    source: dict,
) -> float | None:
    confidence = source.get("confidence") or {}

    values = [
        float(value)
        for value in confidence.values()
        if isinstance(value, (int, float))
    ]

    if not values:
        return None

    return round(sum(values) / len(values), 3)


@app.post(
    "/api/paper-capture/broilers/extract",
    response_model=BroilerPaperCaptureExtractOut,
)
async def extract_broiler_paper_capture(
    company_id: int | None = Form(None),
    image: UploadFile = File(...),
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    content_type = (
        image.content_type
        or "application/octet-stream"
    ).lower()

    allowed_types = {
        "image/jpeg",
        "image/png",
        "image/webp",
    }

    if content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=(
                "Upload a JPG, PNG or WEBP photo of the completed sheet."
            ),
        )

    image_bytes = await image.read()

    if not image_bytes:
        raise HTTPException(
            status_code=400,
            detail="The uploaded image is empty.",
        )

    if len(image_bytes) > 12 * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail="The uploaded image must be 12 MB or smaller.",
        )

    source = _paper_capture_extract_with_openai(
        image_bytes,
        content_type,
    )

    template_id = str(
        source.get("template_id")
        or ""
    ).strip().upper()

    template_parts = _paper_capture_template_parts(
        template_id
    )

    if template_parts is None:
        raise HTTPException(
            status_code=422,
            detail=(
                "OviCore could not read a valid template ID from the sheet. "
                "Retake the photo with the top-right template area visible."
            ),
        )

    plan_id, entry_date = template_parts

    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id == plan_id,
            BroilerPlacementPlan.company_id
            == resolved_company_id,
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail=(
                "The sheet template does not match an OviCore "
                "broiler cycle for this company."
            ),
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    proposed, warnings = _paper_capture_proposed(
        source
    )

    confidence = _paper_capture_overall_confidence(
        source
    )

    low_confidence_fields = [
        field
        for field, value in (
            source.get("confidence") or {}
        ).items()
        if isinstance(value, (int, float))
        and value < 0.80
    ]

    if low_confidence_fields:
        warnings.append(
            "Low-confidence fields require review: "
            + ", ".join(
                field.replace("_", " ")
                for field in low_confidence_fields
            )
            + "."
        )

    capture = BroilerPaperCapture(
        company_id=resolved_company_id,
        placement_plan_id=plan.id,
        template_id=template_id,
        template_version=PAPER_CAPTURE_TEMPLATE_VERSION,
        entry_date=entry_date,
        source_filename=image.filename,
        source_mime_type=content_type,
        source_image_base64=base64.b64encode(
            image_bytes
        ).decode("ascii"),
        raw_extraction_json=json.dumps(
            source,
            ensure_ascii=False,
        ),
        reviewed_json=json.dumps(
            proposed,
            ensure_ascii=False,
        ),
        overall_confidence=confidence,
        status="Review Required",
        extracted_by_model=PAPER_CAPTURE_MODEL,
    )

    db.add(capture)
    db.commit()
    db.refresh(capture)

    age_days = (
        (entry_date - plan.placement_date).days
        if plan.placement_date
        else None
    )

    return BroilerPaperCaptureExtractOut(
        id=capture.id,
        company_id=capture.company_id,
        placement_plan_id=capture.placement_plan_id,
        template_id=capture.template_id,
        entry_date=capture.entry_date,
        farm_name=(
            plan.farm.farm_name
            if plan.farm
            else None
        ),
        shed_name=(
            plan.shed.shed_name
            if plan.shed
            else None
        ),
        cycle_code=plan.cycle_code,
        age_days=age_days,
        status=capture.status,
        overall_confidence=capture.overall_confidence,
        source=BroilerPaperCaptureSourceData.model_validate(
            source
        ),
        proposed=BroilerPaperCaptureReview.model_validate(
            proposed
        ),
        warnings=warnings,
    )


@app.post(
    "/api/paper-capture/broilers/{capture_id}/approve",
    response_model=BroilerPaperCaptureApproveOut,
)
def approve_broiler_paper_capture(
    capture_id: int,
    payload: BroilerPaperCaptureApproveIn,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    capture = (
        db.query(BroilerPaperCapture)
        .filter(
            BroilerPaperCapture.id == capture_id
        )
        .first()
    )

    if not capture:
        raise HTTPException(
            status_code=404,
            detail="Paper Capture record not found.",
        )

    plan = (
        db.query(BroilerPlacementPlan)
        .options(
            joinedload(BroilerPlacementPlan.farm),
            joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerPlacementPlan.id
            == capture.placement_plan_id
        )
        .first()
    )

    if not plan:
        raise HTTPException(
            status_code=404,
            detail="The linked Broiler cycle no longer exists.",
        )

    require_farm_access(
        db,
        current_user,
        plan.farm_id,
    )

    if (
        not current_user.is_global_admin
        and capture.company_id
        != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company.",
        )

    existing = (
        db.query(BroilerDailyPerformance)
        .filter(
            BroilerDailyPerformance.company_id
            == capture.company_id,
            BroilerDailyPerformance.placement_plan_id
            == capture.placement_plan_id,
            BroilerDailyPerformance.entry_date
            == capture.entry_date,
        )
        .first()
    )

    if existing:
        raise HTTPException(
            status_code=409,
            detail=(
                "Daily Data Entry already contains a record for this "
                "cycle and date. Review the existing entry instead of "
                "overwriting it from paper."
            ),
        )

    reviewed = payload.reviewed

    age_days = (
        (capture.entry_date - plan.placement_date).days
        if plan.placement_date
        else None
    )

    entry = BroilerDailyPerformance(
        company_id=capture.company_id,
        placement_plan_id=capture.placement_plan_id,
        entry_date=capture.entry_date,
        age_days=age_days,
        opening_birds=reviewed.opening_birds,

        mortality_front=reviewed.mortality_front,
        mortality_middle=reviewed.mortality_middle,
        mortality_back=reviewed.mortality_back,
        mortality_other=reviewed.mortality_other,

        cull_legs=reviewed.cull_legs,
        cull_runts=reviewed.cull_runts,
        cull_beak=reviewed.cull_beak,
        cull_other=reviewed.cull_other,

        feed_kg=reviewed.feed_kg,
        water_litres=reviewed.water_litres,
        avg_weight_kg=reviewed.body_weight_kg,
        notes=reviewed.notes,

        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    recalculate_daily_performance_entry(entry)

    db.add(entry)
    db.flush()

    capture.performance_entry_id = entry.id
    capture.reviewed_json = json.dumps(
        reviewed.model_dump(),
        ensure_ascii=False,
    )
    capture.status = "Approved"
    capture.reviewed_by = current_user.full_name
    capture.reviewed_at = datetime.utcnow()

    db.commit()
    db.refresh(entry)

    entry = (
        db.query(BroilerDailyPerformance)
        .options(
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.farm),
            joinedload(
                BroilerDailyPerformance.placement_plan
            ).joinedload(BroilerPlacementPlan.shed),
        )
        .filter(
            BroilerDailyPerformance.id == entry.id
        )
        .first()
    )

    return BroilerPaperCaptureApproveOut(
        capture_id=capture.id,
        performance_entry=build_daily_performance_response(
            entry,
            cumulative_mortality_birds=(
                entry.mortality_birds or 0
            ),
        ),
        status=capture.status,
    )




def _import_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _import_bool(value, default: bool = True) -> bool:
    if value is None or _import_text(value) == "":
        return default
    return _import_text(value).lower() in {
        "1", "true", "yes", "y", "on", "active"
    }


def _import_int(
    value,
    field_name: str,
    row_number: int,
    errors: list[str],
) -> int | None:
    if value is None or _import_text(value) == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        errors.append(
            f"Row {row_number}: {field_name} must be a whole number."
        )
        return None


def _import_float(
    value,
    field_name: str,
    row_number: int,
    errors: list[str],
) -> float | None:
    if value is None or _import_text(value) == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        errors.append(
            f"Row {row_number}: {field_name} must be numeric."
        )
        return None


def _import_date(
    value,
    field_name: str,
    row_number: int,
    errors: list[str],
) -> date | None:
    if value is None or _import_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _import_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    errors.append(
        f"Row {row_number}: {field_name} must be a valid date, "
        "preferably yyyy-mm-dd."
    )
    return None


def _sheet_records(
    workbook,
    sheet_name: str,
    required_headers: set[str] | None = None,
) -> list[tuple[int, dict[str, object]]]:
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    header_row: int | None = None
    headers: list[str] = []
    records: list[tuple[int, dict[str, object]]] = []
    required = required_headers or set()

    for row_number, cells in enumerate(sheet.iter_rows(), start=1):
        values = [cell.value for cell in cells]

        if header_row is None:
            if row_number > 20:
                return []

            normalised = [
                _import_text(value).rstrip(" *").strip()
                for value in values
            ]
            present = {value for value in normalised if value}

            if (
                (required and required.issubset(present))
                or (
                    not required
                    and any(
                        _import_text(value).endswith("*")
                        for value in values
                    )
                )
            ):
                header_row = row_number
                headers = normalised
            continue

        if not any(
            value is not None and _import_text(value) != ""
            for value in values
        ):
            continue

        record = {
            headers[index]: (
                values[index] if index < len(values) else None
            )
            for index in range(len(headers))
            if headers[index]
        }
        records.append((row_number, record))

    return records


@app.post("/api/admin/data-import")
async def import_master_data(
    company_id: int = Form(...),
    commit: bool = Form(False),
    allow_updates: bool = Form(False),
    workbook: UploadFile = File(...),
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_global_admin:
        raise HTTPException(
            status_code=403,
            detail="Global Admin access required",
        )

    company = (
        db.query(Company)
        .filter(
            Company.id == company_id,
            Company.active == True,
        )
        .first()
    )

    if not company:
        raise HTTPException(
            status_code=404,
            detail="Selected company was not found or is inactive.",
        )

    filename = workbook.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Upload an .xlsx OviCore import workbook.",
        )

    try:
        content = await workbook.read()
        excel = load_workbook(
            filename=BytesIO(content),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Could not read workbook: {exc}",
        )

    farm_rows = _sheet_records(
        excel,
        "Farms",
        {"Farm Code", "Farm Name", "Active"},
    )
    shed_rows = _sheet_records(
        excel,
        "Sheds",
        {"Farm Code", "Shed Code", "Shed Name"},
    )
    flock_rows = _sheet_records(
        excel,
        "Flocks",
        {"Farm Code", "Shed Code", "Flock Code"},
    )
    standard_rows = _sheet_records(
        excel,
        "Breed Standard",
        {"Standard Code", "Breed", "Age Day", "Bodyweight g"},
    )
    performance_rows = _sheet_records(
        excel,
        "Daily Performance",
        {"Farm Code", "Shed Code", "Flock Code", "Entry Date"},
    )

    errors: list[str] = []
    warnings: list[str] = []
    actions = {
        "farms": {"create": 0, "update": 0, "unchanged": 0},
        "sheds": {"create": 0, "update": 0, "unchanged": 0},
        "flocks": {"create": 0, "update": 0, "unchanged": 0},
        "standards": {"create": 0, "update": 0, "unchanged": 0},
        "performance": {"create": 0, "update": 0, "unchanged": 0},
    }

    if not farm_rows:
        errors.append(
            "The Farms sheet is missing or contains no data rows."
        )
    if not shed_rows:
        errors.append(
            "The Sheds sheet is missing or contains no data rows."
        )
    if not flock_rows:
        errors.append(
            "The Flocks sheet is missing or contains no data rows."
        )
    if not standard_rows:
        warnings.append(
            "The Breed Standard sheet is missing or contains no data rows."
        )
    if not performance_rows:
        warnings.append(
            "The Daily Performance sheet is missing or contains no data rows."
        )

    existing_farms = (
        db.query(BroilerFarm)
        .filter(BroilerFarm.company_id == company_id)
        .all()
    )
    farm_by_code = {
        _import_text(farm.farm_code).lower(): farm
        for farm in existing_farms
        if _import_text(farm.farm_code)
    }

    workbook_farm_codes: set[str] = set()
    parsed_farms: list[dict[str, object]] = []

    for row_number, row in farm_rows:
        farm_code = _import_text(row.get("Farm Code"))
        farm_name = _import_text(row.get("Farm Name"))

        if not farm_code:
            errors.append(
                f"Farms row {row_number}: Farm Code is required."
            )
            continue
        if not farm_name:
            errors.append(
                f"Farms row {row_number}: Farm Name is required."
            )
            continue

        code_key = farm_code.lower()
        if code_key in workbook_farm_codes:
            errors.append(
                f"Farms row {row_number}: duplicate Farm Code "
                f"'{farm_code}'."
            )
            continue
        workbook_farm_codes.add(code_key)

        active = _import_bool(row.get("Active"), True)
        existing = farm_by_code.get(code_key)

        if existing is None:
            actions["farms"]["create"] += 1
        else:
            changed = (
                existing.farm_name != farm_name
                or bool(existing.active) != active
            )
            if changed and allow_updates:
                actions["farms"]["update"] += 1
            elif changed:
                actions["farms"]["unchanged"] += 1
                warnings.append(
                    f"Farms row {row_number}: '{farm_code}' exists "
                    "and differs; enable Allow updates to change it."
                )
            else:
                actions["farms"]["unchanged"] += 1

        parsed_farms.append({
            "row": row_number,
            "farm_code": farm_code,
            "farm_name": farm_name,
            "active": active,
            "existing": existing,
        })

    existing_sheds = (
        db.query(BroilerShed)
        .filter(BroilerShed.company_id == company_id)
        .all()
    )
    farm_code_by_id = {
        farm.id: _import_text(farm.farm_code).lower()
        for farm in existing_farms
    }
    shed_by_key = {
        (
            farm_code_by_id.get(shed.farm_id, ""),
            _import_text(shed.shed_code).lower(),
        ): shed
        for shed in existing_sheds
        if _import_text(shed.shed_code)
    }

    workbook_shed_keys: set[tuple[str, str]] = set()
    parsed_sheds: list[dict[str, object]] = []

    for row_number, row in shed_rows:
        farm_code = _import_text(row.get("Farm Code"))
        shed_code = _import_text(row.get("Shed Code"))
        shed_name = _import_text(row.get("Shed Name"))

        if not farm_code or not shed_code or not shed_name:
            errors.append(
                f"Sheds row {row_number}: Farm Code, Shed Code and "
                "Shed Name are required."
            )
            continue

        composite_key = (farm_code.lower(), shed_code.lower())

        if (
            composite_key[0] not in workbook_farm_codes
            and composite_key[0] not in farm_by_code
        ):
            errors.append(
                f"Sheds row {row_number}: Farm Code "
                f"'{farm_code}' was not found."
            )
            continue

        if composite_key in workbook_shed_keys:
            errors.append(
                f"Sheds row {row_number}: duplicate Shed Code "
                f"'{shed_code}' for Farm Code '{farm_code}'."
            )
            continue
        workbook_shed_keys.add(composite_key)

        floor_area = _import_float(
            row.get("Floor Area m²"),
            "Floor Area m²",
            row_number,
            errors,
        )
        active = _import_bool(row.get("Active"), True)
        existing = shed_by_key.get(composite_key)

        if existing is None:
            actions["sheds"]["create"] += 1
        else:
            changed = (
                existing.shed_name != shed_name
                or bool(existing.active) != active
                or (
                    floor_area is not None
                    and float(existing.floor_area_m2 or 0) != floor_area
                )
            )
            if changed and allow_updates:
                actions["sheds"]["update"] += 1
            elif changed:
                actions["sheds"]["unchanged"] += 1
                warnings.append(
                    f"Sheds row {row_number}: "
                    f"'{farm_code}/{shed_code}' exists and differs; "
                    "enable Allow updates to change it."
                )
            else:
                actions["sheds"]["unchanged"] += 1

        parsed_sheds.append({
            "row": row_number,
            "farm_code": farm_code,
            "shed_code": shed_code,
            "shed_name": shed_name,
            "floor_area_m2": floor_area or 1.0,
            "active": active,
            "existing": existing,
        })

    existing_plans = (
        db.query(BroilerPlacementPlan)
        .filter(BroilerPlacementPlan.company_id == company_id)
        .all()
    )
    plan_by_code = {
        _import_text(plan.cycle_code).lower(): plan
        for plan in existing_plans
        if _import_text(plan.cycle_code)
    }

    workbook_flock_codes: set[str] = set()
    parsed_flocks: list[dict[str, object]] = []

    for row_number, row in flock_rows:
        farm_code = _import_text(row.get("Farm Code"))
        shed_code = _import_text(row.get("Shed Code"))
        flock_code = _import_text(row.get("Flock Code"))
        module = _import_text(row.get("Module"))

        if not farm_code or not shed_code or not flock_code:
            errors.append(
                f"Flocks row {row_number}: Farm Code, Shed Code and "
                "Flock Code are required."
            )
            continue

        if module and module.lower() != "broilers":
            warnings.append(
                f"Flocks row {row_number}: module '{module}' was "
                "skipped; this importer currently creates broiler "
                "flocks only."
            )
            continue

        shed_lookup_key = (farm_code.lower(), shed_code.lower())
        if (
            shed_lookup_key not in workbook_shed_keys
            and shed_lookup_key not in shed_by_key
        ):
            errors.append(
                f"Flocks row {row_number}: shed "
                f"'{farm_code}/{shed_code}' was not found."
            )
            continue

        code_key = flock_code.lower()
        if code_key in workbook_flock_codes:
            errors.append(
                f"Flocks row {row_number}: duplicate Flock Code "
                f"'{flock_code}'."
            )
            continue
        workbook_flock_codes.add(code_key)

        placement_date = _import_date(
            row.get("Placement Date"),
            "Placement Date",
            row_number,
            errors,
        )
        placed_birds = _import_int(
            row.get("Placed Birds"),
            "Placed Birds",
            row_number,
            errors,
        )
        if placement_date is None or placed_birds is None:
            continue
        if placed_birds <= 0:
            errors.append(
                f"Flocks row {row_number}: Placed Birds must be "
                "greater than zero."
            )
            continue

        planned_end = _import_date(
            row.get("Planned Processing/Transfer Date"),
            "Planned Processing/Transfer Date",
            row_number,
            errors,
        )
        growout_days = (
            max(1, (planned_end - placement_date).days)
            if planned_end
            else 42
        )
        status = _import_text(row.get("Status")) or "Active"
        notes = _import_text(row.get("Notes"))
        existing = plan_by_code.get(code_key)

        if existing is None:
            actions["flocks"]["create"] += 1
        else:
            changed = (
                existing.placement_date != placement_date
                or int(existing.planned_birds or 0) != placed_birds
                or _import_text(existing.status) != status
                or _import_text(existing.notes) != notes
            )
            if changed and allow_updates:
                actions["flocks"]["update"] += 1
            elif changed:
                actions["flocks"]["unchanged"] += 1
                warnings.append(
                    f"Flocks row {row_number}: '{flock_code}' exists "
                    "and differs; enable Allow updates to change it."
                )
            else:
                actions["flocks"]["unchanged"] += 1

        parsed_flocks.append({
            "row": row_number,
            "farm_code": farm_code,
            "shed_code": shed_code,
            "flock_code": flock_code,
            "breed": _import_text(row.get("Breed")) or None,
            "placement_date": placement_date,
            "placed_birds": placed_birds,
            "growout_days": growout_days,
            "status": status,
            "notes": notes,
            "existing": existing,
        })

    parsed_standards: list[dict[str, object]] = []
    standard_codes: set[str] = set()

    for row_number, row in standard_rows:
        standard_code = (
            _import_text(row.get("Standard Code"))
            .upper()
            .replace(" ", "_")
        )
        breed = _import_text(row.get("Breed"))
        age_day = _import_int(
            row.get("Age Day"),
            "Age Day",
            row_number,
            errors,
        )
        bodyweight_g = _import_float(
            row.get("Bodyweight g"),
            "Bodyweight g",
            row_number,
            errors,
        )
        daily_feed = _import_float(
            row.get("Daily Feed g/bird"),
            "Daily Feed g/bird",
            row_number,
            errors,
        )
        livability = _import_float(
            row.get("Target Livability %"),
            "Target Livability %",
            row_number,
            errors,
        )

        if not standard_code or not breed:
            errors.append(
                f"Breed Standard row {row_number}: Standard Code "
                "and Breed are required."
            )
            continue
        if age_day is None or bodyweight_g is None:
            continue

        standard_codes.add(standard_code)
        parsed_standards.append({
            "row": row_number,
            "standard_code": standard_code,
            "standard_name": f"{breed} Broiler Standard",
            "breed": breed,
            "age_day": age_day,
            "body_weight_g": bodyweight_g,
            "feed_avg_g_bird_day": daily_feed,
            "liveability_pct": livability,
        })

    if len(standard_codes) > 1:
        errors.append(
            "The Breed Standard sheet must contain only one "
            "Standard Code per workbook."
        )

    existing_standard_rows: list[PerformanceStandard] = []
    active_standard_code = (
        next(iter(standard_codes)) if standard_codes else None
    )

    if active_standard_code:
        existing_standard_rows = (
            db.query(PerformanceStandard)
            .filter(
                PerformanceStandard.standard_code
                == active_standard_code,
                PerformanceStandard.standard_type == "Breed",
                PerformanceStandard.company_id.is_(None),
            )
            .all()
        )

        if not existing_standard_rows:
            actions["standards"]["create"] = len(parsed_standards)
        elif allow_updates:
            actions["standards"]["update"] = len(parsed_standards)
        else:
            actions["standards"]["unchanged"] = len(
                existing_standard_rows
            )
            warnings.append(
                f"Breed standard '{active_standard_code}' already "
                "exists. Enable Allow updates to replace it."
            )

    existing_performance = (
        db.query(BroilerDailyPerformance)
        .filter(BroilerDailyPerformance.company_id == company_id)
        .all()
    )
    perf_by_key = {
        (entry.placement_plan_id, entry.entry_date): entry
        for entry in existing_performance
    }
    parsed_performance: list[dict[str, object]] = []
    workbook_perf_keys: set[tuple[str, date]] = set()

    for row_number, row in performance_rows:
        flock_code = _import_text(row.get("Flock Code"))
        entry_date = _import_date(
            row.get("Entry Date"),
            "Entry Date",
            row_number,
            errors,
        )

        if not flock_code or entry_date is None:
            errors.append(
                f"Daily Performance row {row_number}: Flock Code "
                "and Entry Date are required."
            )
            continue

        code_key = flock_code.lower()
        if (
            code_key not in workbook_flock_codes
            and code_key not in plan_by_code
        ):
            errors.append(
                f"Daily Performance row {row_number}: Flock Code "
                f"'{flock_code}' was not found."
            )
            continue

        workbook_key = (code_key, entry_date)
        if workbook_key in workbook_perf_keys:
            errors.append(
                f"Daily Performance row {row_number}: duplicate "
                f"entry for '{flock_code}' on {entry_date}."
            )
            continue
        workbook_perf_keys.add(workbook_key)

        age_days = _import_int(
            row.get("Age Days"),
            "Age Days",
            row_number,
            errors,
        )
        opening = _import_int(
            row.get("Opening Birds"),
            "Opening Birds",
            row_number,
            errors,
        )
        mortality = _import_int(
            row.get("Mortality Birds"),
            "Mortality Birds",
            row_number,
            errors,
        ) or 0
        culls = _import_int(
            row.get("Cull Birds"),
            "Cull Birds",
            row_number,
            errors,
        ) or 0
        closing = _import_int(
            row.get("Closing Birds"),
            "Closing Birds",
            row_number,
            errors,
        )
        feed_kg = _import_float(
            row.get("Feed kg"),
            "Feed kg",
            row_number,
            errors,
        )
        water_l = _import_float(
            row.get("Water L"),
            "Water L",
            row_number,
            errors,
        )
        bodyweight = _import_float(
            row.get("Bodyweight kg"),
            "Bodyweight kg",
            row_number,
            errors,
        )
        notes = _import_text(row.get("Comments"))

        existing_plan = plan_by_code.get(code_key)
        existing_entry = (
            perf_by_key.get((existing_plan.id, entry_date))
            if existing_plan
            else None
        )

        if existing_entry is None:
            actions["performance"]["create"] += 1
        else:
            changed = (
                int(existing_entry.opening_birds or 0)
                != int(opening or 0)
                or int(existing_entry.mortality_birds or 0)
                != mortality
                or int(existing_entry.cull_birds or 0) != culls
                or float(existing_entry.feed_kg or 0)
                != float(feed_kg or 0)
                or float(existing_entry.water_litres or 0)
                != float(water_l or 0)
                or float(existing_entry.avg_weight_kg or 0)
                != float(bodyweight or 0)
            )
            if changed and allow_updates:
                actions["performance"]["update"] += 1
            else:
                actions["performance"]["unchanged"] += 1

        parsed_performance.append({
            "row": row_number,
            "flock_code": flock_code,
            "entry_date": entry_date,
            "age_days": age_days,
            "opening_birds": opening,
            "mortality_birds": mortality,
            "cull_birds": culls,
            "closing_birds": closing,
            "feed_kg": feed_kg,
            "water_litres": water_l,
            "avg_weight_kg": bodyweight,
            "notes": notes,
            "existing": existing_entry,
        })

    result = {
        "company": {
            "id": company.id,
            "name": (
                getattr(company, "company_name", None)
                or getattr(company, "name", None)
                or f"Company {company.id}"
            ),
        },
        "filename": filename,
        "mode": "commit" if commit else "preview",
        "allow_updates": allow_updates,
        "actions": actions,
        "errors": errors,
        "warnings": warnings,
        "committed": False,
    }

    if errors or not commit:
        return result

    try:
        farm_objects: dict[str, BroilerFarm] = dict(farm_by_code)

        for item in parsed_farms:
            existing = item["existing"]
            if existing is None:
                existing = BroilerFarm(
                    company_id=company_id,
                    farm_code=item["farm_code"],
                    farm_name=item["farm_name"],
                    active=item["active"],
                )
                db.add(existing)
                db.flush()
            elif allow_updates:
                existing.farm_name = item["farm_name"]
                existing.active = item["active"]

            farm_objects[str(item["farm_code"]).lower()] = existing

        shed_objects: dict[
            tuple[str, str],
            BroilerShed,
        ] = dict(shed_by_key)

        for item in parsed_sheds:
            farm = farm_objects[str(item["farm_code"]).lower()]
            key = (
                str(item["farm_code"]).lower(),
                str(item["shed_code"]).lower(),
            )
            existing = item["existing"]

            if existing is None:
                existing = BroilerShed(
                    company_id=company_id,
                    farm_id=farm.id,
                    shed_code=item["shed_code"],
                    shed_name=item["shed_name"],
                    floor_area_m2=item["floor_area_m2"],
                    default_density_kg_m2=32.0,
                    default_target_lw_kg=2.5,
                    default_growout_days=42,
                    active=item["active"],
                )
                db.add(existing)
                db.flush()
            elif allow_updates:
                existing.farm_id = farm.id
                existing.shed_name = item["shed_name"]
                existing.floor_area_m2 = item["floor_area_m2"]
                existing.active = item["active"]

            shed_objects[key] = existing

        plan_objects: dict[str, BroilerPlacementPlan] = dict(
            plan_by_code
        )

        for item in parsed_flocks:
            key = (
                str(item["farm_code"]).lower(),
                str(item["shed_code"]).lower(),
            )
            farm = farm_objects[str(item["farm_code"]).lower()]
            shed = shed_objects[key]
            existing = item["existing"]

            if existing is None:
                existing = BroilerPlacementPlan(
                    company_id=company_id,
                    farm_id=farm.id,
                    shed_id=shed.id,
                    cycle_code=item["flock_code"],
                    placement_date=item["placement_date"],
                    planned_birds=item["placed_birds"],
                    target_density_kg_m2=(
                        shed.default_density_kg_m2
                    ),
                    target_lw_kg=shed.default_target_lw_kg,
                    growout_days=item["growout_days"],
                    chick_allowance_pct=1.5,
                    notes=item["notes"],
                    status=item["status"],
                    last_saved_by=current_user.full_name,
                    last_saved_at=datetime.utcnow(),
                )
                db.add(existing)
                db.flush()
            elif allow_updates:
                existing.farm_id = farm.id
                existing.shed_id = shed.id
                existing.placement_date = item["placement_date"]
                existing.planned_birds = item["placed_birds"]
                existing.growout_days = item["growout_days"]
                existing.notes = item["notes"]
                existing.status = item["status"]
                existing.last_saved_by = current_user.full_name
                existing.last_saved_at = datetime.utcnow()

            plan_objects[
                str(item["flock_code"]).lower()
            ] = existing

        if parsed_standards and (
            not existing_standard_rows or allow_updates
        ):
            if existing_standard_rows:
                (
                    db.query(PerformanceStandard)
                    .filter(
                        PerformanceStandard.standard_code
                        == active_standard_code,
                        PerformanceStandard.standard_type == "Breed",
                        PerformanceStandard.company_id.is_(None),
                    )
                    .delete(synchronize_session=False)
                )

            now = datetime.utcnow()
            for item in parsed_standards:
                db.add(
                    PerformanceStandard(
                        company_id=None,
                        standard_code=item["standard_code"],
                        standard_name=item["standard_name"],
                        standard_type="Breed",
                        module="Broilers",
                        species="Chicken",
                        breed=item["breed"],
                        phase="Growout",
                        age_day=item["age_day"],
                        age_week=None,
                        body_weight_g=item["body_weight_g"],
                        feed_avg_g_bird_day=(
                            item["feed_avg_g_bird_day"]
                        ),
                        liveability_pct=item["liveability_pct"],
                        source_file=filename,
                        active=True,
                        created_at=now,
                        updated_at=now,
                        imported_by=current_user.full_name,
                    )
                )

        for item in parsed_performance:
            plan = plan_objects[
                str(item["flock_code"]).lower()
            ]
            existing = item["existing"]

            if existing is None:
                existing = BroilerDailyPerformance(
                    company_id=company_id,
                    placement_plan_id=plan.id,
                    entry_date=item["entry_date"],
                )
                db.add(existing)
            elif not allow_updates:
                continue

            existing.age_days = item["age_days"]
            existing.opening_birds = item["opening_birds"]
            existing.mortality_front = 0
            existing.mortality_middle = 0
            existing.mortality_back = 0
            existing.mortality_other = item["mortality_birds"]
            existing.cull_legs = 0
            existing.cull_runts = 0
            existing.cull_beak = 0
            existing.cull_other = item["cull_birds"]
            existing.closing_birds = item["closing_birds"]
            existing.feed_kg = item["feed_kg"]
            existing.water_litres = item["water_litres"]
            existing.avg_weight_kg = item["avg_weight_kg"]
            existing.notes = item["notes"]
            existing.last_saved_by = current_user.full_name
            existing.last_saved_at = datetime.utcnow()
            recalculate_daily_performance_entry(existing)

        db.commit()
        result["committed"] = True
        return result

    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Import failed and was rolled back: {exc}",
        )
# ---------------------------------------------------------------------
# Commercial Rearing Flock Register
# ---------------------------------------------------------------------


def _layer_rearing_age_weeks(
    flock: models.LayerRearingFlock,
) -> float | None:
    """
    Calculate flock age using hatch date where available.

    Placement date is used as a fallback so older records without a
    hatch date can still display an age.
    """
    start_date = flock.hatch_date or flock.placement_date

    if start_date is None:
        return None

    return round((date.today() - start_date).days / 7, 2)


def _layer_rearing_days_to_transfer(
    flock: models.LayerRearingFlock,
) -> int | None:
    if flock.planned_transfer_date is None:
        return None

    return (flock.planned_transfer_date - date.today()).days


def _layer_rearing_transfer_readiness(
    flock: models.LayerRearingFlock,
    days_to_transfer: int | None,
) -> str:
    status = (flock.status or "").strip().lower()

    if status == "transferred":
        return "Transferred"

    if status == "closed":
        return "Closed"

    if days_to_transfer is None:
        return "Not assessed"

    if days_to_transfer < 0:
        return "Review required"

    if days_to_transfer <= 7:
        return "Ready for transfer"

    if days_to_transfer <= 14:
        return "Review required"

    return "Not assessed"


def build_layer_rearing_flock_response(
    flock: models.LayerRearingFlock,
) -> LayerRearingFlockOut:
    days_to_transfer = _layer_rearing_days_to_transfer(flock)

    return LayerRearingFlockOut(
        id=flock.id,
        company_id=flock.company_id,
        farm_id=flock.farm_id,
        shed_id=flock.shed_id,
        farm_name=(
            flock.farm.farm_name
            if flock.farm is not None
            else ""
        ),
        shed_name=(
            flock.shed.shed_name
            if flock.shed is not None
            else ""
        ),
        destination_farm_id=flock.destination_farm_id,
        destination_shed_id=flock.destination_shed_id,
        destination_farm_name=(
            flock.destination_farm.farm_name
            if flock.destination_farm is not None
            else None
        ),
        destination_shed_name=(
            flock.destination_shed.shed_name
            if flock.destination_shed is not None
            else None
        ),
        flock_code=flock.flock_code,
        breed=flock.breed,
        hatch_date=flock.hatch_date,
        placement_date=flock.placement_date,
        birds_placed=flock.birds_placed,
        planned_transfer_date=flock.planned_transfer_date,
        actual_transfer_date=flock.actual_transfer_date,
        birds_transferred=flock.birds_transferred,
        transfer_notes=flock.transfer_notes,
        transferred_by=flock.transferred_by,
        transferred_at=flock.transferred_at,
        commercial_layer_flock_id=(
            flock.commercial_layer_flock.id
            if getattr(flock, "commercial_layer_flock", None)
            else None
        ),
        current_age_weeks=_layer_rearing_age_weeks(flock),

        # These fields will later be driven by the shared Daily House Card.
        # Until daily rearing performance is connected, birds placed is the
        # best available opening position and performance variances remain
        # unassessed.
        current_birds=flock.birds_placed,
        cumulative_mortality_pct=None,
        bodyweight_variance_pct=None,

        days_to_transfer=days_to_transfer,
        transfer_readiness=_layer_rearing_transfer_readiness(
            flock,
            days_to_transfer,
        ),
        status=flock.status,
        notes=flock.notes,
        last_saved_by=flock.last_saved_by,
        last_saved_at=flock.last_saved_at,
    )


def _get_layer_rearing_flock(
    db: Session,
    flock_id: int,
) -> models.LayerRearingFlock:
    flock = (
        db.query(models.LayerRearingFlock)
        .options(
            joinedload(models.LayerRearingFlock.farm),
            joinedload(models.LayerRearingFlock.shed),
            joinedload(models.LayerRearingFlock.destination_farm),
            joinedload(models.LayerRearingFlock.destination_shed),
            joinedload(models.LayerRearingFlock.commercial_layer_flock),
        )
        .filter(models.LayerRearingFlock.id == flock_id)
        .first()
    )

    if flock is None:
        raise HTTPException(
            status_code=404,
            detail="Commercial Rearing flock not found",
        )

    return flock


def _validate_layer_rearing_location(
    db: Session,
    current_user: models.AppUser,
    company_id: int,
    farm_id: int,
    shed_id: int,
    *,
    require_user_farm_access: bool,
    location_label: str,
) -> tuple[BroilerFarm, BroilerShed]:
    farm = (
        db.query(BroilerFarm)
        .filter(
            BroilerFarm.id == farm_id,
            BroilerFarm.company_id == company_id,
            BroilerFarm.active == True,
        )
        .first()
    )

    if farm is None:
        raise HTTPException(
            status_code=404,
            detail=f"{location_label} farm was not found or is inactive",
        )

    if (
        require_user_farm_access
        and not access.user_has_farm_access(
            db,
            current_user,
            farm.id,
        )
    ):
        raise HTTPException(
            status_code=403,
            detail=f"You do not have access to the {location_label.lower()} farm",
        )

    shed = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.id == shed_id,
            BroilerShed.company_id == company_id,
            BroilerShed.farm_id == farm.id,
            BroilerShed.active == True,
        )
        .first()
    )

    if shed is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"The selected {location_label.lower()} shed does not "
                "belong to the selected farm and company, or is inactive."
            ),
        )

    return farm, shed


def _validate_layer_rearing_destination(
    db: Session,
    current_user: models.AppUser,
    company_id: int,
    destination_farm_id: int | None,
    destination_shed_id: int | None,
) -> tuple[int | None, int | None]:
    if destination_farm_id is None and destination_shed_id is None:
        return None, None

    if destination_farm_id is None or destination_shed_id is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Destination farm and destination shed must either both "
                "be selected or both be blank."
            ),
        )

    _validate_layer_rearing_location(
        db,
        current_user,
        company_id,
        destination_farm_id,
        destination_shed_id,
        require_user_farm_access=False,
        location_label="Destination",
    )

    return destination_farm_id, destination_shed_id


@app.get(
    "/api/layers/rearing/flocks",
    response_model=list[LayerRearingFlockOut],
)
def list_layer_rearing_flocks(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(models.LayerRearingFlock)
        .options(
            joinedload(models.LayerRearingFlock.farm),
            joinedload(models.LayerRearingFlock.shed),
            joinedload(models.LayerRearingFlock.destination_farm),
            joinedload(models.LayerRearingFlock.destination_shed),
        )
        .filter(
            models.LayerRearingFlock.company_id
            == resolved_company_id
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            models.LayerRearingFlock.farm_id.in_(
                permitted_farm_ids
            )
        )

    flocks = (
        query
        .order_by(
            models.LayerRearingFlock.placement_date.desc(),
            models.LayerRearingFlock.id.desc(),
        )
        .all()
    )

    return [
        build_layer_rearing_flock_response(flock)
        for flock in flocks
    ]


@app.post(
    "/api/layers/rearing/flocks",
    response_model=LayerRearingFlockOut,
)
def create_layer_rearing_flock(
    payload: LayerRearingFlockCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        payload.company_id,
    )

    flock_code = payload.flock_code.strip()

    if not flock_code:
        raise HTTPException(
            status_code=400,
            detail="Flock code is required",
        )

    _validate_layer_rearing_location(
        db,
        current_user,
        resolved_company_id,
        payload.farm_id,
        payload.shed_id,
        require_user_farm_access=True,
        location_label="Rearing",
    )

    destination_farm_id, destination_shed_id = (
        _validate_layer_rearing_destination(
            db,
            current_user,
            resolved_company_id,
            payload.destination_farm_id,
            payload.destination_shed_id,
        )
    )

    duplicate = (
        db.query(models.LayerRearingFlock)
        .filter(
            models.LayerRearingFlock.company_id
            == resolved_company_id,
            models.LayerRearingFlock.flock_code
            == flock_code,
        )
        .first()
    )

    if duplicate is not None:
        raise HTTPException(
            status_code=400,
            detail="Flock code already exists for this company",
        )

    flock = models.LayerRearingFlock(
        company_id=resolved_company_id,
        farm_id=payload.farm_id,
        shed_id=payload.shed_id,
        destination_farm_id=destination_farm_id,
        destination_shed_id=destination_shed_id,
        flock_code=flock_code,
        breed=payload.breed.strip() if payload.breed else None,
        hatch_date=payload.hatch_date,
        placement_date=payload.placement_date,
        birds_placed=payload.birds_placed,
        planned_transfer_date=payload.planned_transfer_date,
        status=payload.status.strip() or "Draft",
        notes=payload.notes.strip() if payload.notes else "",
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(flock)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Flock code already exists for this company",
        )

    return build_layer_rearing_flock_response(
        _get_layer_rearing_flock(db, flock.id)
    )


@app.post(
    "/api/layers/rearing/flocks/new-row",
    response_model=LayerRearingFlockOut,
)
def create_layer_rearing_flock_new_row(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerShed)
        .join(
            BroilerFarm,
            BroilerFarm.id == BroilerShed.farm_id,
        )
        .filter(
            BroilerShed.company_id == resolved_company_id,
            BroilerShed.active == True,
            BroilerFarm.company_id == resolved_company_id,
            BroilerFarm.active == True,
            BroilerFarm.farm_type == "layer_rearing",
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerShed.farm_id.in_(permitted_farm_ids)
        )

    shed = (
        query
        .order_by(
            BroilerFarm.farm_name.asc(),
            BroilerShed.shed_name.asc(),
        )
        .first()
    )

    if shed is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "No active Commercial Rearing sheds are available for "
                "this user and company. Create or classify a farm as "
                "'layer_rearing' first."
            ),
        )

    existing_count = (
        db.query(models.LayerRearingFlock)
        .filter(
            models.LayerRearingFlock.company_id
            == resolved_company_id
        )
        .count()
    )

    next_number = existing_count + 1
    flock_code = f"LR-NEW-{next_number:03d}"

    while (
        db.query(models.LayerRearingFlock)
        .filter(
            models.LayerRearingFlock.company_id
            == resolved_company_id,
            models.LayerRearingFlock.flock_code
            == flock_code,
        )
        .first()
        is not None
    ):
        next_number += 1
        flock_code = f"LR-NEW-{next_number:03d}"

    flock = models.LayerRearingFlock(
        company_id=resolved_company_id,
        farm_id=shed.farm_id,
        shed_id=shed.id,
        destination_farm_id=None,
        destination_shed_id=None,
        flock_code=flock_code,
        breed=None,
        hatch_date=None,
        placement_date=None,
        birds_placed=None,
        planned_transfer_date=None,
        status="Draft",
        notes="",
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(flock)
    db.commit()

    return build_layer_rearing_flock_response(
        _get_layer_rearing_flock(db, flock.id)
    )


@app.patch(
    "/api/layers/rearing/flocks/{flock_id}",
    response_model=LayerRearingFlockOut,
)
def update_layer_rearing_flock(
    flock_id: int,
    payload: LayerRearingFlockPatch,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_layer_rearing_flock(db, flock_id)

    if (
        not current_user.is_global_admin
        and flock.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this rearing farm",
        )

    data = payload.model_dump(exclude_unset=True)

    target_farm_id = data.get("farm_id", flock.farm_id)
    target_shed_id = data.get("shed_id", flock.shed_id)

    _validate_layer_rearing_location(
        db,
        current_user,
        flock.company_id,
        target_farm_id,
        target_shed_id,
        require_user_farm_access=True,
        location_label="Rearing",
    )

    destination_farm_id = data.get(
        "destination_farm_id",
        flock.destination_farm_id,
    )
    destination_shed_id = data.get(
        "destination_shed_id",
        flock.destination_shed_id,
    )

    destination_farm_id, destination_shed_id = (
        _validate_layer_rearing_destination(
            db,
            current_user,
            flock.company_id,
            destination_farm_id,
            destination_shed_id,
        )
    )

    if "flock_code" in data:
        flock_code = (data["flock_code"] or "").strip()

        if not flock_code:
            raise HTTPException(
                status_code=400,
                detail="Flock code is required",
            )

        duplicate = (
            db.query(models.LayerRearingFlock)
            .filter(
                models.LayerRearingFlock.company_id
                == flock.company_id,
                models.LayerRearingFlock.flock_code
                == flock_code,
                models.LayerRearingFlock.id != flock.id,
            )
            .first()
        )

        if duplicate is not None:
            raise HTTPException(
                status_code=400,
                detail="Flock code already exists for this company",
            )

        data["flock_code"] = flock_code

    if "breed" in data:
        data["breed"] = (
            data["breed"].strip()
            if data["breed"]
            else None
        )

    if "status" in data:
        if (data["status"] or "").strip().lower() == "transferred":
            raise HTTPException(
                status_code=400,
                detail=(
                    "Use Transfer selected flock to complete a transfer. "
                    "The status cannot be manually changed to Transferred."
                ),
            )

        data["status"] = (
            data["status"].strip()
            if data["status"]
            else "Draft"
        )

    if "notes" in data:
        data["notes"] = (
            data["notes"].strip()
            if data["notes"]
            else ""
        )

    data["farm_id"] = target_farm_id
    data["shed_id"] = target_shed_id
    data["destination_farm_id"] = destination_farm_id
    data["destination_shed_id"] = destination_shed_id

    for field, value in data.items():
        setattr(flock, field, value)

    flock.last_saved_by = current_user.full_name
    flock.last_saved_at = datetime.utcnow()

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Flock code already exists for this company",
        )

    return build_layer_rearing_flock_response(
        _get_layer_rearing_flock(db, flock.id)
    )


@app.delete("/api/layers/rearing/flocks/{flock_id}")
def delete_layer_rearing_flock(
    flock_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_layer_rearing_flock(db, flock_id)

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    if (
        not current_user.is_global_admin
        and flock.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this rearing farm",
        )

    if (flock.status or "").strip().lower() not in {
        "draft",
        "planned",
    }:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Planned flocks can be deleted. "
                "Use Transferred or Closed for operational history."
            ),
        )

    deleted_id = flock.id
    deleted_code = flock.flock_code

    db.delete(flock)
    db.commit()

    return {
        "deleted": True,
        "id": deleted_id,
        "flock_code": deleted_code,
    }


# ---------------------------------------------------------------------
# Breeder Rearing Flock Register
# ---------------------------------------------------------------------

def _get_breeder_rearing_flock(db: Session, flock_id: int):
    flock = (
        db.query(models.BreederRearingFlock)
        .options(
            joinedload(models.BreederRearingFlock.farm),
            joinedload(models.BreederRearingFlock.shed),
            joinedload(models.BreederRearingFlock.destination_farm),
            joinedload(models.BreederRearingFlock.destination_shed),
            joinedload(models.BreederRearingFlock.production_flock),
        )
        .filter(models.BreederRearingFlock.id == flock_id)
        .first()
    )
    if flock is None:
        raise HTTPException(status_code=404, detail="Breeder Rearing flock not found")
    return flock

def _validate_breeder_rearing_location(
    db: Session, current_user: models.AppUser, company_id: int,
    farm_id: int, shed_id: int, required_type: str, require_access: bool, label: str,
):
    farm = (
        db.query(BroilerFarm)
        .filter(
            BroilerFarm.id == farm_id,
            BroilerFarm.company_id == company_id,
            BroilerFarm.active == True,
            BroilerFarm.farm_type == required_type,
        )
        .first()
    )
    if farm is None:
        raise HTTPException(status_code=400, detail=f"{label} farm has the wrong farm type or is inactive")
    if require_access and not access.user_has_farm_access(db, current_user, farm.id):
        raise HTTPException(status_code=403, detail=f"You do not have access to the {label.lower()} farm")
    shed = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.id == shed_id,
            BroilerShed.company_id == company_id,
            BroilerShed.farm_id == farm.id,
            BroilerShed.active == True,
        )
        .first()
    )
    if shed is None:
        raise HTTPException(status_code=400, detail=f"The selected {label.lower()} shed does not belong to the selected farm")
    return farm, shed

def _breeder_rearing_response(flock):
    females = int(flock.female_birds or 0)
    males = int(flock.male_birds or 0)
    start_date = flock.hatch_date or flock.placement_date
    age_weeks = round((date.today() - start_date).days / 7, 2) if start_date else None
    days_to_transfer = (flock.planned_transfer_date - date.today()).days if flock.planned_transfer_date else None
    male_ratio = round((males / females) * 100, 2) if females > 0 else None
    return BreederRearingFlockOut(
        id=flock.id, company_id=flock.company_id,
        farm_id=flock.farm_id, shed_id=flock.shed_id,
        farm_name=flock.farm.farm_name if flock.farm else "",
        shed_name=flock.shed.shed_name if flock.shed else "",
        destination_farm_id=flock.destination_farm_id,
        destination_shed_id=flock.destination_shed_id,
        destination_farm_name=flock.destination_farm.farm_name if flock.destination_farm else None,
        destination_shed_name=flock.destination_shed.shed_name if flock.destination_shed else None,
        flock_code=flock.flock_code, breed=flock.breed,
        hatch_date=flock.hatch_date, placement_date=flock.placement_date,
        female_birds=flock.female_birds, male_birds=flock.male_birds,
        total_birds=females + males, male_ratio_pct=male_ratio,
        planned_transfer_date=flock.planned_transfer_date,
        actual_transfer_date=flock.actual_transfer_date,
        females_transferred=flock.females_transferred,
        males_transferred=flock.males_transferred,
        transfer_notes=flock.transfer_notes,
        transferred_by=flock.transferred_by,
        transferred_at=flock.transferred_at,
        production_flock_id=(
            flock.production_flock.id
            if getattr(flock, "production_flock", None)
            else None
        ),
        current_age_weeks=age_weeks, days_to_transfer=days_to_transfer,
        status=flock.status, notes=flock.notes,
        last_saved_by=flock.last_saved_by, last_saved_at=flock.last_saved_at,
    )

@app.get("/api/breeders/rearing/flocks", response_model=list[BreederRearingFlockOut])
def list_breeder_rearing_flocks(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(current_user, company_id)
    query = (
        db.query(models.BreederRearingFlock)
        .options(
            joinedload(models.BreederRearingFlock.farm),
            joinedload(models.BreederRearingFlock.shed),
            joinedload(models.BreederRearingFlock.destination_farm),
            joinedload(models.BreederRearingFlock.destination_shed),
            joinedload(models.BreederRearingFlock.production_flock),
        )
        .filter(models.BreederRearingFlock.company_id == resolved_company_id)
    )
    if not (current_user.is_global_admin or current_user.is_company_admin):
        permitted = db.query(models.UserFarmAccess.farm_id).filter(models.UserFarmAccess.user_id == current_user.id)
        query = query.filter(models.BreederRearingFlock.farm_id.in_(permitted))
    return [_breeder_rearing_response(row) for row in query.order_by(models.BreederRearingFlock.id.desc()).all()]

@app.post("/api/breeders/rearing/flocks/new-row", response_model=BreederRearingFlockOut)
def create_breeder_rearing_flock_new_row(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(current_user, company_id)
    query = (
        db.query(BroilerShed)
        .join(BroilerFarm, BroilerFarm.id == BroilerShed.farm_id)
        .filter(
            BroilerShed.company_id == resolved_company_id,
            BroilerShed.active == True,
            BroilerFarm.company_id == resolved_company_id,
            BroilerFarm.active == True,
            BroilerFarm.farm_type == "breeder_rearing",
        )
    )
    if not (current_user.is_global_admin or current_user.is_company_admin):
        permitted = db.query(models.UserFarmAccess.farm_id).filter(models.UserFarmAccess.user_id == current_user.id)
        query = query.filter(BroilerShed.farm_id.in_(permitted))
    shed = query.order_by(BroilerFarm.farm_name.asc(), BroilerShed.shed_name.asc()).first()
    if shed is None:
        raise HTTPException(status_code=400, detail="No active Breeder Rearing shed is available")
    next_number = db.query(models.BreederRearingFlock).filter(models.BreederRearingFlock.company_id == resolved_company_id).count() + 1
    flock_code = f"BRR-NEW-{next_number:03d}"
    while db.query(models.BreederRearingFlock).filter(models.BreederRearingFlock.company_id == resolved_company_id, models.BreederRearingFlock.flock_code == flock_code).first():
        next_number += 1
        flock_code = f"BRR-NEW-{next_number:03d}"
    flock = models.BreederRearingFlock(
        company_id=resolved_company_id, farm_id=shed.farm_id, shed_id=shed.id,
        flock_code=flock_code, status="Draft", notes="",
        last_saved_by=current_user.full_name, last_saved_at=datetime.utcnow(),
    )
    db.add(flock)
    db.commit()
    return _breeder_rearing_response(_get_breeder_rearing_flock(db, flock.id))

@app.patch("/api/breeders/rearing/flocks/{flock_id}", response_model=BreederRearingFlockOut)
def update_breeder_rearing_flock(
    flock_id: int, payload: BreederRearingFlockPatch,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_breeder_rearing_flock(db, flock_id)
    if not current_user.is_global_admin and flock.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="You do not have access to this company")
    if not access.user_has_farm_access(db, current_user, flock.farm_id):
        raise HTTPException(status_code=403, detail="You do not have access to this Breeder Rearing farm")
    data = payload.model_dump(exclude_unset=True)
    farm_id = data.get("farm_id", flock.farm_id)
    shed_id = data.get("shed_id", flock.shed_id)
    _validate_breeder_rearing_location(db, current_user, flock.company_id, farm_id, shed_id, "breeder_rearing", True, "Breeder Rearing")
    destination_farm_id = data.get("destination_farm_id", flock.destination_farm_id)
    destination_shed_id = data.get("destination_shed_id", flock.destination_shed_id)
    if (destination_farm_id is None) != (destination_shed_id is None):
        raise HTTPException(status_code=400, detail="Destination farm and shed must both be selected or both be blank")
    if destination_farm_id is not None:
        _validate_breeder_rearing_location(db, current_user, flock.company_id, destination_farm_id, destination_shed_id, "breeder_layers", False, "Breeder Production")
    if "flock_code" in data:
        code = (data["flock_code"] or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="Flock code is required")
        duplicate = db.query(models.BreederRearingFlock).filter(models.BreederRearingFlock.company_id == flock.company_id, models.BreederRearingFlock.flock_code == code, models.BreederRearingFlock.id != flock.id).first()
        if duplicate:
            raise HTTPException(status_code=400, detail="Flock code already exists")
        data["flock_code"] = code
    if "status" in data and (data["status"] or "").strip().lower() == "transferred":
        raise HTTPException(
            status_code=400,
            detail=(
                "Use Transfer selected flock to complete a transfer. "
                "The status cannot be manually changed to Transferred."
            ),
        )

    for field in ("breed", "status", "notes"):
        if field in data:
            data[field] = data[field].strip() if data[field] else ("Draft" if field == "status" else None if field == "breed" else "")
    data.update({"farm_id": farm_id, "shed_id": shed_id, "destination_farm_id": destination_farm_id, "destination_shed_id": destination_shed_id})
    for field, value in data.items():
        setattr(flock, field, value)
    flock.last_saved_by = current_user.full_name
    flock.last_saved_at = datetime.utcnow()
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Flock code already exists")
    return _breeder_rearing_response(_get_breeder_rearing_flock(db, flock.id))

@app.delete("/api/breeders/rearing/flocks/{flock_id}")
def delete_breeder_rearing_flock(
    flock_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_breeder_rearing_flock(db, flock_id)
    if not (current_user.is_global_admin or current_user.is_company_admin):
        raise HTTPException(status_code=403, detail="Admin access required")
    if (flock.status or "").strip().lower() not in {"draft", "planned"}:
        raise HTTPException(status_code=400, detail="Only Draft or Planned flocks can be deleted")
    deleted_id = flock.id
    db.delete(flock)
    db.commit()
    return {"deleted": True, "id": deleted_id}


# ---------------------------------------------------------------------
# Breeder Rearing -> Breeder Production Transfer
# ---------------------------------------------------------------------

def _breeder_production_response(
    flock: models.BreederProductionFlock,
) -> BreederProductionFlockOut:
    females = int(flock.opening_female_birds or 0)
    males = int(flock.opening_male_birds or 0)

    return BreederProductionFlockOut(
        id=flock.id,
        company_id=flock.company_id,
        source_rearing_flock_id=flock.source_rearing_flock_id,
        farm_id=flock.farm_id,
        shed_id=flock.shed_id,
        farm_name=flock.farm.farm_name if flock.farm else "",
        shed_name=flock.shed.shed_name if flock.shed else "",
        flock_code=flock.flock_code,
        breed=flock.breed,
        hatch_date=flock.hatch_date,
        transfer_date=flock.transfer_date,
        opening_female_birds=females,
        opening_male_birds=males,
        total_opening_birds=females + males,
        male_ratio_pct=(
            round((males / females) * 100, 2)
            if females > 0
            else None
        ),
        status=flock.status,
        notes=flock.notes,
        last_saved_by=flock.last_saved_by,
        last_saved_at=flock.last_saved_at,
    )


def _get_breeder_production_flock(
    db: Session,
    flock_id: int,
) -> models.BreederProductionFlock:
    flock = (
        db.query(models.BreederProductionFlock)
        .options(
            joinedload(models.BreederProductionFlock.farm),
            joinedload(models.BreederProductionFlock.shed),
        )
        .filter(models.BreederProductionFlock.id == flock_id)
        .first()
    )

    if flock is None:
        raise HTTPException(
            status_code=404,
            detail="Breeder Production flock not found",
        )

    return flock


@app.get(
    "/api/breeders/production/flocks",
    response_model=list[BreederProductionFlockOut],
)
def list_breeder_production_flocks(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(models.BreederProductionFlock)
        .options(
            joinedload(models.BreederProductionFlock.farm),
            joinedload(models.BreederProductionFlock.shed),
        )
        .filter(
            models.BreederProductionFlock.company_id
            == resolved_company_id
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            models.BreederProductionFlock.farm_id.in_(
                permitted_farm_ids
            )
        )

    flocks = (
        query
        .order_by(
            models.BreederProductionFlock.transfer_date.desc(),
            models.BreederProductionFlock.id.desc(),
        )
        .all()
    )

    return [
        _breeder_production_response(flock)
        for flock in flocks
    ]


@app.post(
    "/api/breeders/rearing/flocks/{flock_id}/transfer",
    response_model=BreederTransferResult,
)
def transfer_breeder_rearing_flock(
    flock_id: int,
    payload: BreederRearingTransferCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_breeder_rearing_flock(db, flock_id)

    if (
        not current_user.is_global_admin
        and flock.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this Breeder Rearing farm",
        )

    if (flock.status or "").strip().lower() == "transferred":
        raise HTTPException(
            status_code=409,
            detail="This Breeder Rearing flock has already been transferred",
        )

    existing_production = (
        db.query(models.BreederProductionFlock)
        .filter(
            models.BreederProductionFlock.source_rearing_flock_id
            == flock.id
        )
        .first()
    )

    if existing_production is not None:
        raise HTTPException(
            status_code=409,
            detail="A Breeder Production flock already exists for this rearing flock",
        )

    if payload.females_transferred < 0 or payload.males_transferred < 0:
        raise HTTPException(
            status_code=400,
            detail="Transferred bird numbers cannot be negative",
        )

    available_females = int(flock.female_birds or 0)
    available_males = int(flock.male_birds or 0)

    if payload.females_transferred > available_females:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Females transferred cannot exceed the rearing position "
                f"of {available_females:,}."
            ),
        )

    if payload.males_transferred > available_males:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Males transferred cannot exceed the rearing position "
                f"of {available_males:,}."
            ),
        )

    if (
        payload.females_transferred
        + payload.males_transferred
        <= 0
    ):
        raise HTTPException(
            status_code=400,
            detail="At least one bird must be transferred",
        )

    destination_farm, destination_shed = (
        _validate_breeder_rearing_location(
            db,
            current_user,
            flock.company_id,
            payload.destination_farm_id,
            payload.destination_shed_id,
            "breeder_layers",
            False,
            "Breeder Production",
        )
    )

    production_flock = models.BreederProductionFlock(
        company_id=flock.company_id,
        source_rearing_flock_id=flock.id,
        farm_id=destination_farm.id,
        shed_id=destination_shed.id,
        flock_code=flock.flock_code,
        breed=flock.breed,
        hatch_date=flock.hatch_date,
        transfer_date=payload.actual_transfer_date,
        opening_female_birds=payload.females_transferred,
        opening_male_birds=payload.males_transferred,
        status="Active",
        notes=(
            payload.transfer_notes.strip()
            if payload.transfer_notes
            else ""
        ),
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    flock.destination_farm_id = destination_farm.id
    flock.destination_shed_id = destination_shed.id
    flock.actual_transfer_date = payload.actual_transfer_date
    flock.females_transferred = payload.females_transferred
    flock.males_transferred = payload.males_transferred
    flock.transfer_notes = (
        payload.transfer_notes.strip()
        if payload.transfer_notes
        else ""
    )
    flock.transferred_by = current_user.full_name
    flock.transferred_at = datetime.utcnow()
    flock.status = "Transferred"
    flock.last_saved_by = current_user.full_name
    flock.last_saved_at = datetime.utcnow()

    db.add(production_flock)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "The transfer could not be completed because a matching "
                "Breeder Production flock already exists."
            ),
        )

    db.refresh(production_flock)

    rearing_result = _get_breeder_rearing_flock(
        db,
        flock.id,
    )
    production_result = _get_breeder_production_flock(
        db,
        production_flock.id,
    )

    return BreederTransferResult(
        rearing_flock=_breeder_rearing_response(
            rearing_result
        ),
        production_flock=_breeder_production_response(
            production_result
        ),
    )


# ---------------------------------------------------------------------
# Breeder Production Daily House Card
# ---------------------------------------------------------------------

def _get_breeder_production_daily_entry(
    db: Session,
    entry_id: int,
) -> models.BreederProductionDailyPerformance:
    entry = (
        db.query(models.BreederProductionDailyPerformance)
        .options(
            joinedload(
                models.BreederProductionDailyPerformance.flock
            ).joinedload(models.BreederProductionFlock.farm),
            joinedload(
                models.BreederProductionDailyPerformance.flock
            ).joinedload(models.BreederProductionFlock.shed),
        )
        .filter(
            models.BreederProductionDailyPerformance.id
            == entry_id
        )
        .first()
    )

    if entry is None:
        raise HTTPException(
            status_code=404,
            detail="Breeder Production daily entry not found",
        )

    return entry


def _validate_breeder_production_flock_access(
    db: Session,
    current_user: models.AppUser,
    flock_id: int,
) -> models.BreederProductionFlock:
    flock = _get_breeder_production_flock(db, flock_id)

    if (
        not current_user.is_global_admin
        and flock.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this Breeder Production farm",
        )

    return flock


def _recalculate_breeder_production_entry(
    entry: models.BreederProductionDailyPerformance,
):
    opening_females = int(entry.opening_female_birds or 0)
    opening_males = int(entry.opening_male_birds or 0)

    female_losses = (
        int(entry.female_mortality or 0)
        + int(entry.female_culls or 0)
    )
    male_losses = (
        int(entry.male_mortality or 0)
        + int(entry.male_culls or 0)
    )

    entry.closing_female_birds = max(
        0,
        opening_females - female_losses,
    )
    entry.closing_male_birds = max(
        0,
        opening_males - male_losses,
    )

    return entry


def _recalculate_breeder_production_flock_sequence(
    db: Session,
    flock: models.BreederProductionFlock,
    saved_by: str,
) -> int:
    """
    Recalculate the full Breeder Production bird-position sequence.

    The first daily row opens from the transferred production position.
    Every later row opens from the previous day's closing position.
    Female and male positions are rolled forward independently.
    """
    entries = (
        db.query(models.BreederProductionDailyPerformance)
        .filter(
            models.BreederProductionDailyPerformance.flock_id
            == flock.id
        )
        .order_by(
            models.BreederProductionDailyPerformance.entry_date.asc(),
            models.BreederProductionDailyPerformance.id.asc(),
        )
        .all()
    )

    previous_female_closing = int(
        flock.opening_female_birds or 0
    )
    previous_male_closing = int(
        flock.opening_male_birds or 0
    )

    for entry in entries:
        entry.opening_female_birds = (
            previous_female_closing
        )
        entry.opening_male_birds = (
            previous_male_closing
        )

        _recalculate_breeder_production_entry(entry)

        previous_female_closing = int(
            entry.closing_female_birds or 0
        )
        previous_male_closing = int(
            entry.closing_male_birds or 0
        )

        entry.last_saved_by = saved_by
        entry.last_saved_at = datetime.utcnow()

    return len(entries)


def _breeder_production_daily_response(
    entry: models.BreederProductionDailyPerformance,
) -> BreederProductionDailyPerformanceOut:
    flock = entry.flock
    farm = flock.farm if flock else None
    shed = flock.shed if flock else None

    opening_females = int(entry.opening_female_birds or 0)
    closing_females = int(entry.closing_female_birds or 0)
    closing_males = int(entry.closing_male_birds or 0)
    total_closing = closing_females + closing_males

    total_eggs = int(entry.total_eggs or 0)
    hatching_eggs = int(entry.hatching_eggs or 0)
    floor_eggs = int(entry.floor_eggs or 0)

    production_pct = (
        round((total_eggs / opening_females) * 100, 3)
        if opening_females > 0
        else None
    )

    standard_pct = (
        float(entry.production_standard_pct)
        if entry.production_standard_pct is not None
        else None
    )

    variance_pct = (
        round(production_pct - standard_pct, 3)
        if production_pct is not None
        and standard_pct is not None
        else None
    )

    hatching_egg_pct = (
        round((hatching_eggs / total_eggs) * 100, 3)
        if total_eggs > 0
        else None
    )

    floor_egg_pct = (
        round((floor_eggs / total_eggs) * 100, 3)
        if total_eggs > 0
        else None
    )

    male_ratio_pct = (
        round((closing_males / closing_females) * 100, 3)
        if closing_females > 0
        else None
    )

    feed_kg = (
        float(entry.feed_kg)
        if entry.feed_kg is not None
        else None
    )

    feed_per_bird_g = (
        round((feed_kg * 1000) / total_closing, 3)
        if feed_kg is not None and total_closing > 0
        else None
    )

    return BreederProductionDailyPerformanceOut(
        id=entry.id,
        company_id=entry.company_id,
        flock_id=entry.flock_id,

        farm_name=farm.farm_name if farm else "",
        shed_name=shed.shed_name if shed else "",
        flock_code=flock.flock_code if flock else "",
        breed=flock.breed if flock else None,

        entry_date=entry.entry_date,
        age_days=entry.age_days,

        opening_female_birds=entry.opening_female_birds,
        female_mortality=int(entry.female_mortality or 0),
        female_culls=int(entry.female_culls or 0),
        closing_female_birds=entry.closing_female_birds,

        opening_male_birds=entry.opening_male_birds,
        male_mortality=int(entry.male_mortality or 0),
        male_culls=int(entry.male_culls or 0),
        closing_male_birds=entry.closing_male_birds,

        total_closing_birds=total_closing,
        male_ratio_pct=male_ratio_pct,

        feed_kg=feed_kg,
        water_litres=(
            float(entry.water_litres)
            if entry.water_litres is not None
            else None
        ),
        feed_per_bird_g=feed_per_bird_g,

        female_bodyweight_kg=(
            float(entry.female_bodyweight_kg)
            if entry.female_bodyweight_kg is not None
            else None
        ),
        male_bodyweight_kg=(
            float(entry.male_bodyweight_kg)
            if entry.male_bodyweight_kg is not None
            else None
        ),

        total_eggs=total_eggs,
        hatching_eggs=hatching_eggs,
        floor_eggs=floor_eggs,
        rejects=int(entry.rejects or 0),

        production_pct=production_pct,
        production_standard_pct=standard_pct,
        production_variance_pct=variance_pct,
        hatching_egg_pct=hatching_egg_pct,
        floor_egg_pct=floor_egg_pct,

        notes=entry.notes,
        last_saved_by=entry.last_saved_by,
        last_saved_at=entry.last_saved_at,
    )


@app.get(
    "/api/breeders/production/daily-performance",
    response_model=list[BreederProductionDailyPerformanceOut],
)
def list_breeder_production_daily_performance(
    company_id: int | None = None,
    flock_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(models.BreederProductionDailyPerformance)
        .options(
            joinedload(
                models.BreederProductionDailyPerformance.flock
            ).joinedload(models.BreederProductionFlock.farm),
            joinedload(
                models.BreederProductionDailyPerformance.flock
            ).joinedload(models.BreederProductionFlock.shed),
        )
        .join(
            models.BreederProductionFlock,
            models.BreederProductionFlock.id
            == models.BreederProductionDailyPerformance.flock_id,
        )
        .filter(
            models.BreederProductionDailyPerformance.company_id
            == resolved_company_id,
            models.BreederProductionFlock.company_id
            == resolved_company_id,
        )
    )

    if flock_id is not None:
        flock = _validate_breeder_production_flock_access(
            db,
            current_user,
            flock_id,
        )

        query = query.filter(
            models.BreederProductionDailyPerformance.flock_id
            == flock.id
        )

    elif not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id == current_user.id
            )
        )

        query = query.filter(
            models.BreederProductionFlock.farm_id.in_(
                permitted_farm_ids
            )
        )

    entries = (
        query
        .order_by(
            models.BreederProductionDailyPerformance.entry_date.asc(),
            models.BreederProductionDailyPerformance.id.asc(),
        )
        .all()
    )

    if flock_id is not None:
        flock = _validate_breeder_production_flock_access(
            db,
            current_user,
            flock_id,
        )

        _recalculate_breeder_production_flock_sequence(
            db,
            flock,
            current_user.full_name,
        )

        db.commit()

        entries = (
            query
            .order_by(
                models.BreederProductionDailyPerformance.entry_date.asc(),
                models.BreederProductionDailyPerformance.id.asc(),
            )
            .all()
        )

    return [
        _breeder_production_daily_response(entry)
        for entry in entries
    ]


@app.post(
    "/api/breeders/production/daily-performance/new-row",
    response_model=BreederProductionDailyPerformanceOut,
)
def create_breeder_production_daily_row(
    flock_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _validate_breeder_production_flock_access(
        db,
        current_user,
        flock_id,
    )

    previous = (
        db.query(models.BreederProductionDailyPerformance)
        .filter(
            models.BreederProductionDailyPerformance.flock_id
            == flock.id
        )
        .order_by(
            models.BreederProductionDailyPerformance.entry_date.desc(),
            models.BreederProductionDailyPerformance.id.desc(),
        )
        .first()
    )

    if previous is not None:
        entry_date = previous.entry_date + timedelta(days=1)
        opening_females = int(previous.closing_female_birds or 0)
        opening_males = int(previous.closing_male_birds or 0)
    else:
        entry_date = flock.transfer_date
        opening_females = int(flock.opening_female_birds or 0)
        opening_males = int(flock.opening_male_birds or 0)

    age_days = (
        (entry_date - flock.hatch_date).days
        if flock.hatch_date is not None
        else None
    )

    entry = models.BreederProductionDailyPerformance(
        company_id=flock.company_id,
        flock_id=flock.id,
        entry_date=entry_date,
        age_days=age_days,
        opening_female_birds=opening_females,
        female_mortality=0,
        female_culls=0,
        opening_male_birds=opening_males,
        male_mortality=0,
        male_culls=0,
        feed_kg=None,
        water_litres=None,
        female_bodyweight_kg=None,
        male_bodyweight_kg=None,
        total_eggs=0,
        hatching_eggs=0,
        floor_eggs=0,
        rejects=0,
        production_standard_pct=None,
        notes="",
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    _recalculate_breeder_production_entry(entry)

    db.add(entry)
    db.flush()

    _recalculate_breeder_production_flock_sequence(
        db,
        flock,
        current_user.full_name,
    )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="A daily entry already exists for this flock and date",
        )

    return _breeder_production_daily_response(
        _get_breeder_production_daily_entry(db, entry.id)
    )


@app.post(
    "/api/breeders/production/daily-performance",
    response_model=BreederProductionDailyPerformanceOut,
)
def create_breeder_production_daily_performance(
    payload: BreederProductionDailyPerformanceCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _validate_breeder_production_flock_access(
        db,
        current_user,
        payload.flock_id,
    )

    if payload.company_id != flock.company_id:
        raise HTTPException(
            status_code=400,
            detail="The selected company does not match the flock company",
        )

    existing = (
        db.query(models.BreederProductionDailyPerformance)
        .filter(
            models.BreederProductionDailyPerformance.flock_id
            == flock.id,
            models.BreederProductionDailyPerformance.entry_date
            == payload.entry_date,
        )
        .first()
    )

    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail="A daily entry already exists for this flock and date",
        )

    data = payload.model_dump()
    data["company_id"] = flock.company_id

    entry = models.BreederProductionDailyPerformance(**data)

    _recalculate_breeder_production_entry(entry)

    entry.last_saved_by = current_user.full_name
    entry.last_saved_at = datetime.utcnow()

    db.add(entry)
    db.flush()

    _recalculate_breeder_production_flock_sequence(
        db,
        flock,
        current_user.full_name,
    )

    db.commit()

    return _breeder_production_daily_response(
        _get_breeder_production_daily_entry(db, entry.id)
    )


@app.patch(
    "/api/breeders/production/daily-performance/{entry_id}",
    response_model=BreederProductionDailyPerformanceOut,
)
def update_breeder_production_daily_performance(
    entry_id: int,
    payload: BreederProductionDailyPerformancePatch,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    entry = _get_breeder_production_daily_entry(
        db,
        entry_id,
    )

    flock = _validate_breeder_production_flock_access(
        db,
        current_user,
        entry.flock_id,
    )

    data = payload.model_dump(exclude_unset=True)

    integer_fields = {
        "opening_female_birds",
        "female_mortality",
        "female_culls",
        "opening_male_birds",
        "male_mortality",
        "male_culls",
        "total_eggs",
        "hatching_eggs",
        "floor_eggs",
        "rejects",
    }

    for field in integer_fields:
        if field in data and data[field] is not None and data[field] < 0:
            raise HTTPException(
                status_code=400,
                detail=f"{field.replace('_', ' ').title()} cannot be negative",
            )

    if "entry_date" in data:
        duplicate = (
            db.query(models.BreederProductionDailyPerformance)
            .filter(
                models.BreederProductionDailyPerformance.flock_id
                == flock.id,
                models.BreederProductionDailyPerformance.entry_date
                == data["entry_date"],
                models.BreederProductionDailyPerformance.id
                != entry.id,
            )
            .first()
        )

        if duplicate is not None:
            raise HTTPException(
                status_code=409,
                detail="A daily entry already exists for this flock and date",
            )

        if flock.hatch_date is not None:
            data["age_days"] = (
                data["entry_date"] - flock.hatch_date
            ).days

    for field, value in data.items():
        setattr(entry, field, value)

    _recalculate_breeder_production_entry(entry)

    entry.last_saved_by = current_user.full_name
    entry.last_saved_at = datetime.utcnow()

    _recalculate_breeder_production_flock_sequence(
        db,
        flock,
        current_user.full_name,
    )

    db.commit()

    return _breeder_production_daily_response(
        _get_breeder_production_daily_entry(db, entry.id)
    )


@app.delete(
    "/api/breeders/production/daily-performance/{entry_id}"
)
def delete_breeder_production_daily_performance(
    entry_id: int,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    entry = _get_breeder_production_daily_entry(
        db,
        entry_id,
    )

    _validate_breeder_production_flock_access(
        db,
        current_user,
        entry.flock_id,
    )

    deleted_id = entry.id
    db.delete(entry)
    db.commit()

    return {
        "deleted": True,
        "id": deleted_id,
    }


# ---------------------------------------------------------------------
# Commercial Layer Performance
# ---------------------------------------------------------------------


# ---------------------------------------------------------------------
# Commercial Layer Flock Register CRUD
# ---------------------------------------------------------------------

def _get_commercial_layer_flock(
    db: Session,
    flock_id: int,
) -> models.CommercialLayerFlock:
    flock = (
        db.query(models.CommercialLayerFlock)
        .options(
            joinedload(models.CommercialLayerFlock.farm),
            joinedload(models.CommercialLayerFlock.shed),
            joinedload(models.CommercialLayerFlock.source_rearing_flock),
            joinedload(models.CommercialLayerFlock.daily_performance),
        )
        .filter(
            models.CommercialLayerFlock.id == flock_id
        )
        .first()
    )

    if flock is None:
        raise HTTPException(
            status_code=404,
            detail="Commercial Layer flock not found",
        )

    return flock


def _validate_commercial_layer_location(
    db: Session,
    current_user: models.AppUser,
    company_id: int,
    farm_id: int,
    shed_id: int,
) -> tuple[BroilerFarm, BroilerShed]:
    farm = (
        db.query(BroilerFarm)
        .filter(
            BroilerFarm.id == farm_id,
            BroilerFarm.company_id == company_id,
            BroilerFarm.active == True,
            BroilerFarm.farm_type == "commercial_layers",
        )
        .first()
    )

    if farm is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "The selected farm must be an active "
                "Commercial Layers farm."
            ),
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        farm.id,
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "You do not have access to this "
                "Commercial Layers farm."
            ),
        )

    shed = (
        db.query(BroilerShed)
        .filter(
            BroilerShed.id == shed_id,
            BroilerShed.company_id == company_id,
            BroilerShed.farm_id == farm.id,
            BroilerShed.active == True,
        )
        .first()
    )

    if shed is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "The selected shed does not belong to the "
                "selected Commercial Layers farm and company."
            ),
        )

    return farm, shed


def _commercial_layer_current_values(
    flock: models.CommercialLayerFlock,
) -> dict:
    entries = sorted(
        list(flock.daily_performance or []),
        key=lambda item: (
            item.entry_date or date.min,
            item.id or 0,
        ),
    )

    latest = entries[-1] if entries else None

    cumulative_mortality = sum(
        int(entry.mortality_birds or 0)
        for entry in entries
    )

    housed_birds = int(flock.birds_housed or 0)

    current_birds = (
        int(latest.closing_birds or 0)
        if latest is not None
        else housed_birds
    )

    latest_production_pct = None
    latest_feed_g_bird_day = None

    if latest is not None:
        opening_birds = int(latest.opening_birds or 0)
        total_eggs = int(latest.total_eggs or 0)
        feed_kg = (
            float(latest.feed_kg)
            if latest.feed_kg is not None
            else None
        )

        if opening_birds > 0:
            latest_production_pct = round(
                (total_eggs / opening_birds) * 100,
                3,
            )

        if (
            feed_kg is not None
            and current_birds > 0
        ):
            latest_feed_g_bird_day = round(
                (feed_kg * 1000) / current_birds,
                3,
            )

    cumulative_mortality_pct = (
        round(
            (cumulative_mortality / housed_birds) * 100,
            3,
        )
        if housed_birds > 0
        else None
    )

    start_date = flock.hatch_date or flock.housed_date

    current_age_weeks = (
        round(
            (date.today() - start_date).days / 7,
            2,
        )
        if start_date is not None
        else None
    )

    status = (flock.status or "").strip().lower()

    if status in {"depleted", "closed"}:
        production_status = "Closed"
    elif latest_production_pct is None:
        production_status = "Not started"
    elif latest_production_pct >= 85:
        production_status = "Peak production"
    elif latest_production_pct >= 50:
        production_status = "In production"
    else:
        production_status = "Review"

    return {
        "current_age_weeks": current_age_weeks,
        "current_birds": current_birds,
        "latest_production_pct": latest_production_pct,
        "latest_feed_g_bird_day":
            latest_feed_g_bird_day,
        "cumulative_mortality_pct":
            cumulative_mortality_pct,
        "production_status": production_status,
    }


def _commercial_layer_flock_response(
    flock: models.CommercialLayerFlock,
) -> CommercialLayerFlockOut:
    calculated = _commercial_layer_current_values(
        flock
    )

    return CommercialLayerFlockOut(
        id=flock.id,
        source_rearing_flock_id=(
            flock.source_rearing_flock_id
        ),
        source_rearing_flock_code=(
            flock.source_rearing_flock.flock_code
            if flock.source_rearing_flock
            else None
        ),
        company_id=flock.company_id,
        farm_id=flock.farm_id,
        shed_id=flock.shed_id,
        farm_name=(
            flock.farm.farm_name
            if flock.farm
            else ""
        ),
        shed_name=(
            flock.shed.shed_name
            if flock.shed
            else ""
        ),
        flock_code=flock.flock_code,
        breed=flock.breed,
        hatch_date=flock.hatch_date,
        housed_date=flock.housed_date,
        birds_housed=flock.birds_housed,
        planned_depletion_date=(
            flock.planned_depletion_date
        ),
        current_age_weeks=(
            calculated["current_age_weeks"]
        ),
        current_birds=calculated["current_birds"],
        latest_production_pct=(
            calculated["latest_production_pct"]
        ),
        latest_feed_g_bird_day=(
            calculated["latest_feed_g_bird_day"]
        ),
        cumulative_mortality_pct=(
            calculated["cumulative_mortality_pct"]
        ),
        production_status=(
            calculated["production_status"]
        ),
        status=flock.status,
        notes=flock.notes,
        last_saved_by=flock.last_saved_by,
        last_saved_at=flock.last_saved_at,
    )


@app.post(
    "/api/layers/commercial/flocks/new-row",
    response_model=CommercialLayerFlockOut,
)
def create_commercial_layer_flock_new_row(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(
        get_current_user
    ),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(BroilerShed)
        .join(
            BroilerFarm,
            BroilerFarm.id == BroilerShed.farm_id,
        )
        .filter(
            BroilerShed.company_id
            == resolved_company_id,
            BroilerShed.active == True,
            BroilerFarm.company_id
            == resolved_company_id,
            BroilerFarm.active == True,
            BroilerFarm.farm_type
            == "commercial_layers",
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id
                == current_user.id
            )
        )

        query = query.filter(
            BroilerShed.farm_id.in_(
                permitted_farm_ids
            )
        )

    shed = (
        query
        .order_by(
            BroilerFarm.farm_name.asc(),
            BroilerShed.shed_name.asc(),
        )
        .first()
    )

    if shed is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "No active Commercial Layers shed is "
                "available for this user and company."
            ),
        )

    next_number = (
        db.query(models.CommercialLayerFlock)
        .filter(
            models.CommercialLayerFlock.company_id
            == resolved_company_id
        )
        .count()
        + 1
    )

    flock_code = f"CL-NEW-{next_number:03d}"

    while (
        db.query(models.CommercialLayerFlock)
        .filter(
            models.CommercialLayerFlock.company_id
            == resolved_company_id,
            models.CommercialLayerFlock.flock_code
            == flock_code,
        )
        .first()
        is not None
    ):
        next_number += 1
        flock_code = f"CL-NEW-{next_number:03d}"

    flock = models.CommercialLayerFlock(
        company_id=resolved_company_id,
        source_rearing_flock_id=None,
        farm_id=shed.farm_id,
        shed_id=shed.id,
        flock_code=flock_code,
        breed=None,
        hatch_date=None,
        housed_date=None,
        birds_housed=None,
        planned_depletion_date=None,
        status="Draft",
        notes="",
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(flock)
    db.commit()

    return _commercial_layer_flock_response(
        _get_commercial_layer_flock(
            db,
            flock.id,
        )
    )


@app.post(
    "/api/layers/commercial/flocks",
    response_model=CommercialLayerFlockOut,
)
def create_commercial_layer_flock(
    payload: CommercialLayerFlockCreate,
    current_user: models.AppUser = Depends(
        get_current_user
    ),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        payload.company_id,
    )

    _validate_commercial_layer_location(
        db,
        current_user,
        resolved_company_id,
        payload.farm_id,
        payload.shed_id,
    )

    flock_code = payload.flock_code.strip()

    if not flock_code:
        raise HTTPException(
            status_code=400,
            detail="Flock code is required",
        )

    duplicate = (
        db.query(models.CommercialLayerFlock)
        .filter(
            models.CommercialLayerFlock.company_id
            == resolved_company_id,
            models.CommercialLayerFlock.flock_code
            == flock_code,
        )
        .first()
    )

    if duplicate is not None:
        raise HTTPException(
            status_code=409,
            detail=(
                "Flock code already exists for "
                "this company."
            ),
        )

    flock = models.CommercialLayerFlock(
        company_id=resolved_company_id,
        source_rearing_flock_id=None,
        farm_id=payload.farm_id,
        shed_id=payload.shed_id,
        flock_code=flock_code,
        breed=(
            payload.breed.strip()
            if payload.breed
            else None
        ),
        hatch_date=payload.hatch_date,
        housed_date=payload.housed_date,
        birds_housed=payload.birds_housed,
        planned_depletion_date=(
            payload.planned_depletion_date
        ),
        status=payload.status.strip() or "Draft",
        notes=(
            payload.notes.strip()
            if payload.notes
            else ""
        ),
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    db.add(flock)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "Flock code already exists for "
                "this company."
            ),
        )

    return _commercial_layer_flock_response(
        _get_commercial_layer_flock(
            db,
            flock.id,
        )
    )


@app.patch(
    "/api/layers/commercial/flocks/{flock_id}",
    response_model=CommercialLayerFlockOut,
)
def update_commercial_layer_flock(
    flock_id: int,
    payload: CommercialLayerFlockPatch,
    current_user: models.AppUser = Depends(
        get_current_user
    ),
    db: Session = Depends(get_db),
):
    flock = _get_commercial_layer_flock(
        db,
        flock_id,
    )

    if (
        not current_user.is_global_admin
        and flock.company_id
        != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "You do not have access to "
                "this company."
            ),
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "You do not have access to this "
                "Commercial Layers farm."
            ),
        )

    data = payload.model_dump(
        exclude_unset=True
    )

    target_farm_id = data.get(
        "farm_id",
        flock.farm_id,
    )
    target_shed_id = data.get(
        "shed_id",
        flock.shed_id,
    )

    _validate_commercial_layer_location(
        db,
        current_user,
        flock.company_id,
        target_farm_id,
        target_shed_id,
    )

    if "flock_code" in data:
        flock_code = (
            data["flock_code"] or ""
        ).strip()

        if not flock_code:
            raise HTTPException(
                status_code=400,
                detail="Flock code is required",
            )

        duplicate = (
            db.query(models.CommercialLayerFlock)
            .filter(
                models.CommercialLayerFlock.company_id
                == flock.company_id,
                models.CommercialLayerFlock.flock_code
                == flock_code,
                models.CommercialLayerFlock.id
                != flock.id,
            )
            .first()
        )

        if duplicate is not None:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Flock code already exists for "
                    "this company."
                ),
            )

        data["flock_code"] = flock_code

    if "breed" in data:
        data["breed"] = (
            data["breed"].strip()
            if data["breed"]
            else None
        )

    if "status" in data:
        data["status"] = (
            data["status"].strip()
            if data["status"]
            else "Draft"
        )

    if "notes" in data:
        data["notes"] = (
            data["notes"].strip()
            if data["notes"]
            else ""
        )

    data["farm_id"] = target_farm_id
    data["shed_id"] = target_shed_id

    for field, value in data.items():
        setattr(flock, field, value)

    flock.last_saved_by = current_user.full_name
    flock.last_saved_at = datetime.utcnow()

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "Flock code already exists for "
                "this company."
            ),
        )

    return _commercial_layer_flock_response(
        _get_commercial_layer_flock(
            db,
            flock.id,
        )
    )


@app.delete(
    "/api/layers/commercial/flocks/{flock_id}"
)
def delete_commercial_layer_flock(
    flock_id: int,
    current_user: models.AppUser = Depends(
        get_current_user
    ),
    db: Session = Depends(get_db),
):
    flock = _get_commercial_layer_flock(
        db,
        flock_id,
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        raise HTTPException(
            status_code=403,
            detail="Admin access required",
        )

    if (
        not current_user.is_global_admin
        and flock.company_id
        != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "You do not have access to "
                "this company."
            ),
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "You do not have access to this "
                "Commercial Layers farm."
            ),
        )

    if flock.source_rearing_flock_id is not None:
        raise HTTPException(
            status_code=400,
            detail=(
                "A transferred flock cannot be deleted. "
                "Use Depleted or Closed to preserve the "
                "flock lifecycle."
            ),
        )

    linked_entries = (
        db.query(
            models.CommercialLayerDailyPerformance
        )
        .filter(
            models.CommercialLayerDailyPerformance.flock_id
            == flock.id
        )
        .count()
    )

    if linked_entries > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "This flock has Daily House Card records "
                "and cannot be deleted."
            ),
        )

    if (flock.status or "").strip().lower() not in {
        "draft",
        "planned",
    }:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Planned flocks can be "
                "deleted."
            ),
        )

    deleted_id = flock.id
    deleted_code = flock.flock_code

    db.delete(flock)
    db.commit()

    return {
        "deleted": True,
        "id": deleted_id,
        "flock_code": deleted_code,
    }


@app.get(
    "/api/layers/commercial/flocks",
    response_model=list[CommercialLayerFlockOut],
)
def list_commercial_layer_flocks(
    company_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(models.CommercialLayerFlock)
        .options(
            joinedload(models.CommercialLayerFlock.farm),
            joinedload(models.CommercialLayerFlock.shed),
            joinedload(
                models.CommercialLayerFlock.source_rearing_flock
            ),
            joinedload(
                models.CommercialLayerFlock.daily_performance
            ),
        )
        .filter(
            models.CommercialLayerFlock.company_id
            == resolved_company_id
        )
    )

    if not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id == current_user.id
            )
        )

        query = query.filter(
            models.CommercialLayerFlock.farm_id.in_(
                permitted_farm_ids
            )
        )

    return [
        _commercial_layer_flock_response(flock)
        for flock in (
            query
            .order_by(
                models.CommercialLayerFlock.housed_date.desc(),
                models.CommercialLayerFlock.id.desc(),
            )
            .all()
        )
    ]


@app.get(
    "/api/layers/commercial/performance",
    response_model=list[CommercialLayerPerformanceOut],
)
def list_commercial_layer_performance(
    company_id: int | None = None,
    flock_id: int | None = None,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    resolved_company_id = resolve_company_id(
        current_user,
        company_id,
    )

    query = (
        db.query(models.CommercialLayerDailyPerformance)
        .options(
            joinedload(
                models.CommercialLayerDailyPerformance.flock
            ).joinedload(models.CommercialLayerFlock.farm),
            joinedload(
                models.CommercialLayerDailyPerformance.flock
            ).joinedload(models.CommercialLayerFlock.shed),
        )
        .join(
            models.CommercialLayerFlock,
            models.CommercialLayerFlock.id
            == models.CommercialLayerDailyPerformance.flock_id,
        )
        .filter(
            models.CommercialLayerDailyPerformance.company_id
            == resolved_company_id,
            models.CommercialLayerFlock.company_id
            == resolved_company_id,
        )
    )

    if flock_id is not None:
        flock = (
            db.query(models.CommercialLayerFlock)
            .filter(
                models.CommercialLayerFlock.id == flock_id,
                models.CommercialLayerFlock.company_id
                == resolved_company_id,
            )
            .first()
        )

        if flock is None:
            raise HTTPException(
                status_code=404,
                detail="Commercial Layer flock not found",
            )

        if not access.user_has_farm_access(
            db,
            current_user,
            flock.farm_id,
        ):
            raise HTTPException(
                status_code=403,
                detail="You do not have access to this Commercial Layer farm",
            )

        query = query.filter(
            models.CommercialLayerDailyPerformance.flock_id
            == flock.id
        )

    elif not (
        current_user.is_global_admin
        or current_user.is_company_admin
    ):
        permitted_farm_ids = (
            db.query(models.UserFarmAccess.farm_id)
            .filter(
                models.UserFarmAccess.user_id == current_user.id
            )
        )

        query = query.filter(
            models.CommercialLayerFlock.farm_id.in_(
                permitted_farm_ids
            )
        )

    entries = (
        query
        .order_by(
            models.CommercialLayerDailyPerformance.flock_id.asc(),
            models.CommercialLayerDailyPerformance.entry_date.asc(),
            models.CommercialLayerDailyPerformance.id.asc(),
        )
        .all()
    )

    cumulative_mortality_by_flock: dict[int, int] = {}
    cumulative_eggs_by_flock: dict[int, int] = {}
    output: list[CommercialLayerPerformanceOut] = []

    for entry in entries:
        flock = entry.flock
        flock_id_key = entry.flock_id

        cumulative_mortality_by_flock.setdefault(
            flock_id_key,
            0,
        )
        cumulative_eggs_by_flock.setdefault(
            flock_id_key,
            0,
        )

        cumulative_mortality_by_flock[flock_id_key] += int(
            entry.mortality_birds or 0
        )
        cumulative_eggs_by_flock[flock_id_key] += int(
            entry.total_eggs or 0
        )

        opening_birds = int(entry.opening_birds or 0)
        closing_birds = int(entry.closing_birds or 0)
        total_eggs = int(entry.total_eggs or 0)
        housed_birds = int(flock.birds_housed or 0) if flock else 0
        feed_kg = (
            float(entry.feed_kg)
            if entry.feed_kg is not None
            else None
        )

        production_pct = (
            round((total_eggs / opening_birds) * 100, 3)
            if opening_birds > 0
            else None
        )
        cumulative_mortality_pct = (
            round(
                (
                    cumulative_mortality_by_flock[flock_id_key]
                    / housed_birds
                )
                * 100,
                3,
            )
            if housed_birds > 0
            else None
        )
        feed_g_bird_day = (
            round((feed_kg * 1000) / closing_birds, 3)
            if feed_kg is not None and closing_birds > 0
            else None
        )
        eggs_per_bird_cumulative = (
            round(
                cumulative_eggs_by_flock[flock_id_key]
                / housed_birds,
                4,
            )
            if housed_birds > 0
            else None
        )

        output.append(
            CommercialLayerPerformanceOut(
                id=entry.id,
                company_id=entry.company_id,
                flock_id=entry.flock_id,
                farm_name=(
                    flock.farm.farm_name
                    if flock and flock.farm
                    else ""
                ),
                shed_name=(
                    flock.shed.shed_name
                    if flock and flock.shed
                    else ""
                ),
                flock_code=flock.flock_code if flock else "",
                breed=flock.breed if flock else None,
                entry_date=entry.entry_date,
                age_days=entry.age_days,
                age_weeks=(
                    round(entry.age_days / 7, 2)
                    if entry.age_days is not None
                    else None
                ),
                opening_birds=entry.opening_birds,
                mortality_birds=int(entry.mortality_birds or 0),
                cull_birds=int(entry.cull_birds or 0),
                closing_birds=entry.closing_birds,
                total_eggs=total_eggs,
                production_pct=production_pct,
                cumulative_mortality_pct=cumulative_mortality_pct,
                egg_weight_g=(
                    float(entry.egg_weight_g)
                    if entry.egg_weight_g is not None
                    else None
                ),
                feed_g_bird_day=feed_g_bird_day,
                eggs_per_bird_cumulative=eggs_per_bird_cumulative,
                bodyweight_g=(
                    float(entry.bodyweight_g)
                    if entry.bodyweight_g is not None
                    else None
                ),
                production_standard_pct=(
                    float(entry.production_standard_pct)
                    if entry.production_standard_pct is not None
                    else None
                ),
                mortality_standard_pct=(
                    float(entry.mortality_standard_pct)
                    if entry.mortality_standard_pct is not None
                    else None
                ),
                egg_weight_standard_g=(
                    float(entry.egg_weight_standard_g)
                    if entry.egg_weight_standard_g is not None
                    else None
                ),
                feed_standard_g_bird_day=(
                    float(entry.feed_standard_g_bird_day)
                    if entry.feed_standard_g_bird_day is not None
                    else None
                ),
                eggs_per_bird_standard=(
                    float(entry.eggs_per_bird_standard)
                    if entry.eggs_per_bird_standard is not None
                    else None
                ),
                bodyweight_standard_g=(
                    float(entry.bodyweight_standard_g)
                    if entry.bodyweight_standard_g is not None
                    else None
                ),
                notes=entry.notes,
                last_saved_by=entry.last_saved_by,
                last_saved_at=entry.last_saved_at,
            )
        )

    return output


@app.post(
    "/api/layers/rearing/flocks/{flock_id}/transfer",
    response_model=LayerRearingTransferResult,
)
def transfer_layer_rearing_flock(
    flock_id: int,
    payload: LayerRearingTransferCreate,
    current_user: models.AppUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    flock = _get_layer_rearing_flock(db, flock_id)

    if (
        not current_user.is_global_admin
        and flock.company_id != current_user.company_id
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this company",
        )

    if not access.user_has_farm_access(
        db,
        current_user,
        flock.farm_id,
    ):
        raise HTTPException(
            status_code=403,
            detail="You do not have access to this Commercial Rearing farm",
        )

    if (flock.status or "").strip().lower() == "transferred":
        raise HTTPException(
            status_code=409,
            detail="This Commercial Rearing flock has already been transferred",
        )

    existing = (
        db.query(models.CommercialLayerFlock)
        .filter(
            models.CommercialLayerFlock.source_rearing_flock_id
            == flock.id
        )
        .first()
    )
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail="A Commercial Layer flock already exists for this rearing flock",
        )

    if payload.birds_transferred <= 0:
        raise HTTPException(
            status_code=400,
            detail="Birds transferred must be greater than zero",
        )

    available_birds = int(flock.birds_placed or 0)
    if payload.birds_transferred > available_birds:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Birds transferred cannot exceed the rearing position "
                f"of {available_birds:,}."
            ),
        )

    destination_farm, destination_shed = (
        _validate_layer_rearing_location(
            db,
            current_user,
            flock.company_id,
            payload.destination_farm_id,
            payload.destination_shed_id,
            require_user_farm_access=False,
            location_label="Commercial Layers",
        )
    )

    if destination_farm.farm_type != "commercial_layers":
        raise HTTPException(
            status_code=400,
            detail="The destination farm must be classified as Commercial Layers",
        )

    commercial_flock = models.CommercialLayerFlock(
        source_rearing_flock_id=flock.id,
        company_id=flock.company_id,
        farm_id=destination_farm.id,
        shed_id=destination_shed.id,
        flock_code=flock.flock_code,
        breed=flock.breed,
        hatch_date=flock.hatch_date,
        housed_date=payload.actual_transfer_date,
        birds_housed=payload.birds_transferred,
        status="Active",
        notes=(payload.transfer_notes or "").strip(),
        last_saved_by=current_user.full_name,
        last_saved_at=datetime.utcnow(),
    )

    flock.destination_farm_id = destination_farm.id
    flock.destination_shed_id = destination_shed.id
    flock.actual_transfer_date = payload.actual_transfer_date
    flock.birds_transferred = payload.birds_transferred
    flock.transfer_notes = (payload.transfer_notes or "").strip()
    flock.transferred_by = current_user.full_name
    flock.transferred_at = datetime.utcnow()
    flock.status = "Transferred"
    flock.last_saved_by = current_user.full_name
    flock.last_saved_at = datetime.utcnow()

    db.add(commercial_flock)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "The transfer could not be completed because a matching "
                "Commercial Layer flock already exists."
            ),
        )

    db.refresh(commercial_flock)

    return LayerRearingTransferResult(
        rearing_flock=build_layer_rearing_flock_response(
            _get_layer_rearing_flock(db, flock.id)
        ),
        commercial_layer_flock=_commercial_layer_flock_response(
            (
                db.query(models.CommercialLayerFlock)
                .options(
                    joinedload(models.CommercialLayerFlock.farm),
                    joinedload(models.CommercialLayerFlock.shed),
                )
                .filter(models.CommercialLayerFlock.id == commercial_flock.id)
                .first()
            )
        ),
    )
